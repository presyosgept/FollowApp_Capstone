from django.shortcuts import render

from twilio.rest import Client
from json import encoder
from typing import Counter
from django.db.models.fields import TimeField
from django.http import HttpResponse, request
from django.http.response import Http404, HttpResponsePermanentRedirect, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Max, Min
from django.contrib.auth import authenticate, login, logout

from django.contrib import messages

from django.contrib.auth.decorators import login_required

from .utilities import create_notification, create_feedback
from django.views import generic
from django.utils import timezone
from datetime import date, datetime, timedelta

import datetime as dt
from tablib import Dataset

from django.shortcuts import render
from django.shortcuts import get_object_or_404
import openpyxl

from django.views.generic import View
from django.shortcuts import redirect
from django.contrib import messages
from django.conf import settings
import random
from django.contrib.auth.models import User

from .models import NewTime, Calendar, Subject, School, Department, Faculty, Counselor, SubjectOfferings, DegreeProgram, Student, Studentload
from .models import CounselorFeedback, AccountCreated, StudentAdditionalInformation, Referral, ReferralDetails, Notification, SetScheduleCounselor, NotificationFeedback
from .resources import SubjectResource, SchoolResource, DepartmentResource, FacultyResource, CounselorResource, SubjectOfferingsResource, DegreeProgramResource, StudentResource, StudentloadResource

from .forms import SetActiveForm,StudentSetSchedForm, FilterDateForm, CalendarForm, CounselorFeedbackForm, SetScheduleCounselorForm, ReferralForm, FilterForm, StudentAdditionalInformationForm, EditDegreeProgramForm, EditSchoolForm, CheckSemForm, SearchForm, AssignCounselorForm, CreateUserForm, AccountsForm, VerificationForm, AccountCreatedForm, EditDepartmentForm, EditSubjectForm

# global variables


count = 0
count1 = 0
formm = AccountsForm()
feedback_id = 0
Active_Year = '2022'
Active_Sem = '1st'
global dep
global school
global filterStartDate
global filterEndData
global Search


def register(request):
    form = AccountCreatedForm()
    if request.method == 'POST':
        char = '1234567890'
        for x in range(0, 1):
            code = ''
            for x in range(0, 5):
                code_char = random.choice(char)
                code = code + code_char
        username = request.POST.get('id_number')
        email = request.POST.get('email')
        checker = email.find('+63')
        if checker == 0:
            qs_faculty = Faculty.objects.all()
            qs_student = Student.objects.all()
            qs_acc = AccountCreated.objects.all()

            if username == 'followapp':
                account_sid = 'AC47090e11c4e65aba8e1ce9f75e7522c5'
                auth_token = '2b5813f003934b172e7f429e802e5704'
                client = Client(account_sid, auth_token)
                body = 'This is your VERIFICATION CODE FOR FOLLOWAPP: ' + code
                message = client.messages.create(
                    to=email, from_='+15166045607', body=body)
                print(message.sid)
                value = AccountCreated(
                    id_number=username, email=email, password=code)
                value.save()
                return redirect('verification_code')

            exist = 0
            for acc in qs_acc:
                if username == acc.id_number:
                    exist = 1

            if exist == 0:
                flag = 0
                for user in qs_student:
                    if username == user.student_number:
                        flag = 1
                for user in qs_faculty:
                    if username == user.faculty_id:
                        flag = 1
                if flag == 1:
                    account_sid = 'AC47090e11c4e65aba8e1ce9f75e7522c5'
                    auth_token = '2b5813f003934b172e7f429e802e5704'
                    client = Client(account_sid, auth_token)
                    body = 'This is your VERIFICATION CODE FOR FOLLOWAPP: ' + code
                    message = client.messages.create(
                        to=email, from_='+15166045607', body=body)
                    print(message.sid)
                    value = AccountCreated(
                        id_number=username, email=email, password=code)
                    value.save()
                    messages.info(request, "Check SMS for Code")
                    return redirect('verification_code')
                else:
                    messages.info(request, "Account Not Valid")
            else:
                messages.info(request, "Account Already Existed")
        else:
            messages.info(request, "Number should start +63")
    else:
        AccountsForm()

    return render(request, 'register.html', {'form': form})


def verification_code(request):
    form = VerificationForm()
    code = request.POST.get('code')
    if request.method == 'POST':
        qs_account = AccountCreated.objects.all()
        flag = 0
        for vcode in qs_account:
            if (code == vcode.password):
                flag = 1

        if(flag == 1):
            return redirect('signup')
        else:
            VerificationForm()
            messages.info(request, 'Invalid Code')
    else:
        VerificationForm()

    return render(request, 'verification.html', {'form': form})


def signup(request):
    if request.user.is_authenticated:
        return redirect('login')
    else:
        form = CreateUserForm()
        if request.method == 'POST':
            username = request.POST.get('username')
            qs_account = AccountCreated.objects.all()
            exist = 0
            for user in qs_account:
                if user.id_number == username:
                    exist = 1
            if (exist == 1):
                flag = 0
                qs = Student.objects.all()
                for student in qs:
                    if student.student_number == username:
                        flag = 1
                if flag == 1:
                    form = CreateUserForm(request.POST)
                    if form.is_valid():
                        form.save()
                        user = form.cleaned_data.get('username')
                        messages.info(
                            request, 'Student Account was created for ' + user)
                        return redirect('login')
                elif flag == 0:
                    qs = Faculty.objects.all()
                    for faculty in qs:
                        if faculty.faculty_id == username:
                            form = CreateUserForm(request.POST)
                            if form.is_valid():
                                form.save()
                                user = form.cleaned_data.get('username')
                                messages.info(
                                    request, 'Account was created for ' + user)
                                return redirect('login')

                if username == 'followapp':
                    form = CreateUserForm(request.POST)
                    if form.is_valid():
                        form.save()
                        user = form.cleaned_data.get('username')
                        messages.info(
                            request, 'Admin Account was created for ' + user)
                        return redirect('login')

                messages.info(request, 'Check Credentials Account Not Created')

            else:
                messages.info(request, 'Check Credentials Account Not Created')

    return render(request, 'signup.html', {'form': form})


def loginPage(request):
    if request.user.is_authenticated:
        return redirect('login')
    else:
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')

            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                if request.method == 'POST':
                    flag = 0
                    username = request.POST.get('username')
                    try:
                        qs = Student.objects.get(student_number=username)
                        if qs is not None:
                            flag = 1
                    except Exception:
                        flag = 0
                    if flag == 1:
                        request.session['username'] = username
                        try:
                            stud = StudentAdditionalInformation.objects.get(
                                    student_number=username)
                            check = True
                        except:
                            check = False
                        if check:
                            return redirect('student_home_view')
                        else:
                            return redirect('student_add_information')

                    else:
                        try:
                            qs = Faculty.objects.get(faculty_id=username)
                            if qs is not None:
                                check = True
                        except Exception:
                            check = False
                        if check:
                            if qs.role == 'Teacher':
                                request.session['username'] = username
                                flag = 2
                            elif qs.role == 'Counselor':
                                request.session['username'] = username
                                flag = 3
                            elif qs.role == 'Director':
                                request.session['username'] = username
                                flag = 4
                    if flag == 2:
                        return redirect('teacher_home_view')
                    if flag == 3:
                        return redirect('counselor_home_view')
                    if flag == 4:
                        return redirect('director_home_view')
                    if username == 'followapp':
                        return redirect('admin_home_view')
            else:
                messages.info(request, 'Username or Password is Incorrect')

        return render(request, 'login.html', {})


def logoutUser(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def home(request, *args, **kwargs):
    return render(request, "welcomepage.html", {})


# admin

from django.views.generic import ListView
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.models import User
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger



@login_required(login_url='login')
def admin_home_view(request, *args, **kwargs): 
    global Active_Year
    global Active_Sem     
    return render(request, "admin/home.html", {'Active_Year':Active_Year,'Active_Sem':Active_Sem})

@login_required(login_url='login')
def set_active_year(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    active_form = SetActiveForm(initial={'active_year': '2022'})
    if request.method == "POST":
        active_form = SetActiveForm(request.POST,initial={'active_year': '2022'})
        if active_form.is_valid():
            choice_active_sem = active_form['active_sem'].value()
            choice_active_year = active_form['active_year'].value()
            Active_Sem = choice_active_sem
            Active_Year = choice_active_year
    return render(request, "admin/set_active_year.html", {'active_form':active_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_subject(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_subject = Subject.objects.all().order_by('subject_code')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_subject, 10)
    try:
        subject = paginator.page(page)
    except PageNotAnInteger:
        subject = paginator.page(1)
    except EmptyPage:
        subject = paginator.page(paginator.num_pages)
    return render(request, "admin/view_subject.html", {"subject": subject,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def edit_subject(request, code):
    global Active_Year
    global Active_Sem
    subject = Subject.objects.get(subject_code=code)
    edit_form = EditSubjectForm()
    if request.method == "POST":
        edit_form = EditSubjectForm(request.POST)
        if edit_form.is_valid():
            new_units = edit_form['units'].value()
            edit = Subject.objects.get(subject_code=code)
            edit.units = new_units
            edit.save()
            subject = Subject.objects.get(subject_code=code)
            return redirect('view_subject')
    return render(request, "admin/edit_subject.html", {'subject': subject, 'edit_form': edit_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_school(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_school = School.objects.all().order_by('school_code')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_school, 10)
    try:
        school = paginator.page(page)
    except PageNotAnInteger:
        school = paginator.page(1)
    except EmptyPage:
        school = paginator.page(paginator.num_pages)
    return render(request, "admin/view_school.html", {"school": school,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def edit_school(request, name):
    global Active_Year
    global Active_Sem
    school = School.objects.get(school_name=name)
    edit_form = EditSchoolForm(instance=school)
    if request.method == "POST":
        edit_form = EditSchoolForm(request.POST, instance=school)
        if edit_form.is_valid():
            new_school_code = edit_form['school_code'].value()
            edit = School.objects.get(school_name=name)
            edit.school_code = new_school_code
            edit.save()
            return redirect('view_school')
    return render(request, "admin/edit_school.html", {'school': school, 'edit_form': edit_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_department(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_department = Department.objects.all().order_by('department_code')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_department, 10)
    try:
        department = paginator.page(page)
    except PageNotAnInteger:
        department = paginator.page(1)
    except EmptyPage:
        department = paginator.page(paginator.num_pages)
    return render(request, "admin/view_department.html", {'department': department,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def edit_department(request, code):
    global Active_Year
    global Active_Sem
    department = Department.objects.get(department_code=code)
    edit_form = EditDepartmentForm(instance=department)
    if request.method == "POST":
        edit_form = EditDepartmentForm(request.POST, instance=department)
        if edit_form.is_valid():
            new_department_name = edit_form['department_name'].value()
            new_school_code = edit_form['school_code'].value()
            get_school_code = School.objects.get(
                id=new_school_code)
            edit = Department.objects.get(department_code=code)
            edit.department_name = new_department_name
            edit.school_code = get_school_code
            edit.save()
            return redirect('view_department')
    return render(request, "admin/edit_department.html", {'department': department, 'edit_form': edit_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_faculty(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_faculty = Faculty.objects.all().order_by('lastname')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_faculty, 10)
    try:
        faculty = paginator.page(page)
    except PageNotAnInteger:
        faculty = paginator.page(1)
    except EmptyPage:
        faculty = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_faculty', search=search_choice)
    return render(request, "admin/view_faculty.html", {'faculty': faculty, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def search_faculty(request, search):
    global Active_Year
    global Active_Sem
    all_faculty = Faculty.objects.all().order_by('lastname')
    get_faculty = []
    for obj in all_faculty:
        if search.lower() in obj.lastname.lower():
            get_faculty.append(Faculty(faculty_id=obj.faculty_id,
                                   lastname=obj.lastname, firstname=obj.firstname,
                                   middlename=obj.middlename,
                                   department_code=obj.department_code))
    page = request.GET.get('page', 1)
    paginator = Paginator(get_faculty, 10)
    try:
        faculty = paginator.page(page)
    except PageNotAnInteger:
        faculty = paginator.page(1)
    except EmptyPage:
        faculty = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_faculty', search=search_choice)
    return render(request, "admin/search_faculty.html", {"faculty": faculty, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_faculty_with_load(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_faculty = Faculty.objects.all().order_by('lastname')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_faculty, 10)
    try:
        faculty = paginator.page(page)
    except PageNotAnInteger:
        faculty = paginator.page(1)
    except EmptyPage:
        faculty = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_faculty_with_load', search=search_choice)
    return render(request, "admin/view_faculty_with_load.html", {'faculty': faculty, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def search_faculty_with_load(request, search):
    global Active_Year
    global Active_Sem
    all_faculty = Faculty.objects.all().order_by('lastname')
    get_faculty = []
    for obj in all_faculty:
        if search.lower() in obj.lastname.lower():
            get_faculty.append(Faculty(faculty_id=obj.faculty_id,
                                   lastname=obj.lastname, firstname=obj.firstname,
                                   middlename=obj.middlename,
                                   department_code=obj.department_code))
    page = request.GET.get('page', 1)
    paginator = Paginator(get_faculty, 10)
    try:
        faculty = paginator.page(page)
    except PageNotAnInteger:
        faculty = paginator.page(1)
    except EmptyPage:
        faculty = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_faculty_with_load', search=search_choice)
    return render(request, "admin/search_faculty_with_load.html", {"faculty": faculty, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_faculty_detail(request, faculty_id):
    global Active_Year
    global Active_Sem
    subject_offerings = SubjectOfferings.objects.filter(sem_id=Active_Sem,academic_year=Active_Year)
    get_faculty = Faculty.objects.get(faculty_id=faculty_id)
    get_subject_offerings = []
    for obj in subject_offerings:
        if obj.faculty_id.faculty_id == faculty_id:
            get_subject_offerings.append(SubjectOfferings(
                offer_no=obj.offer_no,
                subject_code=obj.subject_code,
                subject_title=obj.subject_title,
                school_days=obj.school_days,
                school_time=obj.school_time))
    return render(request, "admin/view_faculty_detail.html", {'get_subject_offerings': get_subject_offerings, 'get_faculty': get_faculty,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_counselor(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_counselor = Faculty.objects.filter(role='Counselor').order_by('lastname')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_counselor, 10)
    try:
        counselor = paginator.page(page)
    except PageNotAnInteger:
        counselor = paginator.page(1)
    except EmptyPage:
        counselor = paginator.page(paginator.num_pages)
    return render(request, "admin/view_counselor.html", {'counselor': counselor,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_subject_offerings(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_subject_offerings = SubjectOfferings.objects.all().order_by('offer_no')
    check_sem = SetActiveForm(initial={'active_year': '2022'})
    if request.method == 'POST':
        check_sem = SetActiveForm(request.POST,initial={'active_year': '2022'})
        if check_sem.is_valid():
            sem_choice = check_sem['active_sem'].value()
            sem_year = check_sem['active_year'].value()
            if sem_choice == '--':
                get_subject_offerings = SubjectOfferings.objects.filter(academic_year=sem_year).order_by('offer_no')
            else:
                get_subject_offerings = SubjectOfferings.objects.filter(
                    sem_id=sem_choice,academic_year=sem_year).order_by('offer_no')
    
    page = request.GET.get('page', 1)
    paginator = Paginator(get_subject_offerings, 10)
    try:
        subject_offerings = paginator.page(page)
    except PageNotAnInteger:
        subject_offerings = paginator.page(1)
    except EmptyPage:
        subject_offerings = paginator.page(paginator.num_pages)
    return render(request, "admin/view_subject_offerings.html", {'subject_offerings': subject_offerings, 'check_sem': check_sem,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_degree_program(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_degree_program = DegreeProgram.objects.all().order_by('program_code')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_degree_program, 10)
    try:
        degree_program = paginator.page(page)
    except PageNotAnInteger:
        degree_program = paginator.page(1)
    except EmptyPage:
        degree_program = paginator.page(paginator.num_pages)
    return render(request, "admin/view_degree_program.html", {'degree_program': degree_program,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def edit_degree_program(request, code):
    global Active_Year
    global Active_Sem
    degree_program = DegreeProgram.objects.get(program_code=code)
    edit_form = EditDegreeProgramForm(instance=degree_program)
    if request.method == "POST":
        edit_form = EditDegreeProgramForm(
            request.POST, instance=degree_program)
        if edit_form.is_valid():
            new_program_name = edit_form['program_name'].value()
            new_school_code = edit_form['school_code'].value()
            get_school = School.objects.get(id=new_school_code)
            edit = DegreeProgram.objects.get(program_code=code)
            edit.program_name = new_program_name
            edit.school_code = get_school
            edit.save()
            return redirect('view_degree_program')
    return render(request, "admin/edit_degree_program.html", {'degree_program': degree_program, 'edit_form': edit_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_student(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_student = Student.objects.all().order_by('lastname')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_student, 10)
    try:
        student = paginator.page(page)
    except PageNotAnInteger:
        student = paginator.page(1)
    except EmptyPage:
        student = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student', search=search_choice)
    return render(request, "admin/view_student.html", {'student': student, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def search_student(request, search):
    global Active_Year
    global Active_Sem
    all_student = Student.objects.all().order_by('lastname')
    get_student = []
    for obj in all_student:
        if search.lower() in obj.lastname.lower():
            get_student.append(Student(student_number=obj.student_number,
                                   lastname=obj.lastname, firstname=obj.firstname,
                                   middlename=obj.middlename,  department_code=obj.department_code,
                                   program_code=obj.program_code, academic_year=obj.academic_year,
                                   sem_id=obj.sem_id, school_name=obj.school_name,
                                   student_email=obj.student_email, role=obj.role))
    page = request.GET.get('page', 1)

    paginator = Paginator(get_student, 10)
    try:
        student = paginator.page(page)
    except PageNotAnInteger:
        student = paginator.page(1)
    except EmptyPage:
        student = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student', search=search_choice)
    return render(request, "admin/search_student.html", {"student": student, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_student_with_load(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_student = Student.objects.all().order_by('lastname')
    page = request.GET.get('page', 1)

    paginator = Paginator(get_student, 10)
    try:
        student = paginator.page(page)
    except PageNotAnInteger:
        student = paginator.page(1)
    except EmptyPage:
        student = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student_with_load', search=search_choice)
    return render(request, "admin/view_student_with_load.html", {'student': student, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def search_student_with_load(request, search):
    global Active_Year
    global Active_Sem
    all_student = Student.objects.all().order_by('lastname')
    get_student = []
    for obj in all_student:
        if search.lower() in obj.lastname.lower():
            get_student.append(Student(student_number=obj.student_number,
                                   lastname=obj.lastname, firstname=obj.firstname,
                                   middlename=obj.middlename,  department_code=obj.department_code,
                                   program_code=obj.program_code, academic_year=obj.academic_year,
                                   sem_id=obj.sem_id, school_name=obj.school_name,
                                   student_email=obj.student_email, role=obj.role))
    page = request.GET.get('page', 1)

    paginator = Paginator(get_student, 10)
    try:
        student = paginator.page(page)
    except PageNotAnInteger:
        student = paginator.page(1)
    except EmptyPage:
        student = paginator.page(paginator.num_pages)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student_with_load', search=search_choice)
    return render(request, "admin/search_student_with_load.html", {"student": student, 'search_form': search_form,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def view_student_detail(request, student_number):
    global Active_Year
    global Active_Sem
    subject_offerings = SubjectOfferings.objects.filter(sem_id=Active_Sem,academic_year=Active_Year)
    get_student_load = Studentload.objects.filter(
        student_number=student_number,sem_id=Active_Sem,academic_year=Active_Year)
    get_student = Student.objects.get(student_number=student_number)
    get_degree_name = DegreeProgram.objects.get(program_code = get_student.program_code_id)
    get_subject_offerings = []
    for obj in subject_offerings:
        for obj1 in get_student_load:
            if obj1.offer_no.offer_no == obj.offer_no:
                get_subject_offerings.append(SubjectOfferings(
                    offer_no=obj.offer_no,
                    subject_code=obj.subject_code,
                    subject_title=obj.subject_title,
                    school_days=obj.school_days,
                    school_time=obj.school_time))
    return render(request, "admin/view_student_detail.html", {'get_subject_offerings': get_subject_offerings, 'Active_Year':Active_Year,'Active_Sem':Active_Sem,'get_student': get_student,'get_degree_name':get_degree_name})


@login_required(login_url='login')
def view_student_load(request, *args, **kwargs):
    global Active_Year
    global Active_Sem
    get_student_load = Studentload.objects.filter(sem_id=Active_Sem, academic_year=Active_Year)
    page = request.GET.get('page', 1)
    paginator = Paginator(get_student_load, 10)
    try:
        student_load = paginator.page(page)
    except PageNotAnInteger:
        student_load = paginator.page(1)
    except EmptyPage:
        student_load = paginator.page(paginator.num_pages)
    return render(request, "admin/view_student_load.html", {'student_load': student_load,'Active_Year':Active_Year,'Active_Sem':Active_Sem})

new_faculty_list = []
new_subject_offerings_list=[]
new_student_list=[]
new_student_load_list = []

@login_required(login_url='login')
def upload_faculty(request):
    global Active_Year
    global Active_Sem
    global new_faculty_list
    try:
        get_faculty = Faculty.objects.all()
        if request.method == 'POST':
            FacultyResource()
            dataset = Dataset()
            new_faculty = request.FILES['myfile']
            imported_data = dataset.load(new_faculty.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_faculty)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 7):
                for data in imported_data:
                    new_faculty_list.append({'faculty_id': str(data[0]), 'lastname': data[1],
                                               'firstname': data[2], 'middlename': data[3], 'email': data[4],
                                               'role': data[5],'department_code':data[6]})
                    check_faculty = Faculty.objects.all()
                    flag_faculty = 0
                    flag_depa = 0
                    check_depa = Department.objects.all()
                    for obj in check_depa:
                        if obj.department_code == data[6]:
                            flag_depa = 1
                    if flag_depa == 0:
                        depa = Department(department_code=data[6])
                        depa.save()
                    for obj in check_faculty:
                        if obj.faculty_id == str(data[0]):
                            flag_faculty = 1
                            id = obj.faculty_id
                    if flag_faculty == 0:
                        depa = Department.objects.get(
                            department_code=str(data[6]))
                        value = Faculty(
                            faculty_id=str(data[0]),
                            lastname=data[1],
                            firstname=data[2],
                            middlename=data[3],
                            email=data[4],
                            role=data[5],
                            department_code=depa
                        )
                        value.save()
                    else:
                        depa = Department.objects.get(
                            department_code=str(data[6]))
                        check = Faculty.objects.get(faculty_id=id)
                        check.lastname = data[1]
                        check.firstname = data[2]
                        check.middlename = data[3]
                        check.email = data[4]
                        check.role = data[5]
                        check.department_code = depa
                        check.save()
                        
                        
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
        page = request.GET.get('page', 1)

        paginator = Paginator(new_faculty_list, 10)
        try:
            faculty = paginator.page(page)
        except PageNotAnInteger:
            faculty = paginator.page(1)
        except EmptyPage:
            faculty = paginator.page(paginator.num_pages)
    except Exception:
        faculty=[]
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_faculty.html", {"faculty": faculty,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def upload_subject_offerings(request):
    global Active_Year
    global Active_Sem
    global new_subject_offerings_list
    sem_choice = '1st'
    sem_year='2022'
    try:
        get_subject_offerings = SubjectOfferings.objects.all()
        check_sem = SetActiveForm(initial={'active_year': '2022'})
        if request.method == "POST":
            check_sem = SetActiveForm(request.POST,initial={'active_year': '2022'})
            if check_sem.is_valid():
                sem_choice = check_sem['active_sem'].value()
                sem_year = check_sem['active_year'].value()
            # check_sem = CheckSemForm()
            # if request.method == 'POST':
            #     check_sem = CheckSemForm(request.POST)
            #     if check_sem.is_valid():
            #         sem_choice = check_sem['sem'].value()
            SubjectOfferingsResource()
            dataset = Dataset()
            new_sem = request.FILES['myfile']
            imported_data = dataset.load(new_sem.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_sem)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 9):
                for data in imported_data:
                    check_depa = Department.objects.all()
                    check_faculty = Faculty.objects.all()
                    check_subject = Subject.objects.all()
                    is_depa = bool(check_depa)
                    is_faculty = bool(check_faculty)
                    is_subject = bool(check_subject)

                    if is_depa:
                        try:
                            get_depa_exist = Department.objects.get(
                                department_code=data[7])
                            check_depa = bool(get_depa_exist)
                        except Exception:
                            check_depa = False
                        if check_depa:
                            get_depa_exist.department_code = data[7]
                            get_depa_exist.department_name = get_depa_exist.department_name
                            get_depa_exist.school_code = get_depa_exist.school_code
                            get_depa_exist.save()
                        else:
                            depa = Department(department_code=data[7])
                            depa.save()
                    else:
                        depa = Department(department_code=data[7])
                        depa.save()

                    if is_faculty:
                        depa = Department.objects.get(
                            department_code=data[7])
                        try:
                            get_faculty_exist = Faculty.objects.get(
                                faculty_id=str(data[8]))
                            check_faculty = bool(get_faculty_exist)
                        except Exception:
                            check_faculty = False
                        if check_faculty:
                            get_faculty_exist.faculty_id = get_faculty_exist.faculty_id
                            get_faculty_exist.lastname = get_faculty_exist.lastname
                            get_faculty_exist.firstname = get_faculty_exist.firstname
                            get_faculty_exist.middlename = get_faculty_exist.middlename
                            get_faculty_exist.email = get_faculty_exist.email
                            get_faculty_exist.role = get_faculty_exist.role
                            get_faculty_exist.department_code = depa
                            get_faculty_exist.save()
                        else:
                            faculty = Faculty(faculty_id=str(
                                data[8]), department_code=depa)
                            faculty.save()
                    else:
                        depa = Department.objects.get(
                            department_code=data[7])
                        faculty = Faculty(faculty_id=str(
                            data[8]), department_code=depa)
                        faculty.save()

                    if is_subject:
                        try:
                            get_subject_exist = Subject.objects.get(
                                subject_code=data[1])
                            check_subjects = bool(get_subject_exist)
                        except Exception:
                            check_subjects = False
                        if check_subjects:
                            get_subject_exist.subject_code = data[1]
                            get_subject_exist.subject_title = data[2]
                            get_subject_exist.units = get_subject_exist
                        else:
                            subject = Subject(
                                subject_code=data[1], subject_title=data[2])
                            subject.save()
                    else:
                        subject = Subject(
                            subject_code=data[1], subject_title=data[2])
                        subject.save()

                for data in imported_data:
                    new_subject_offerings_list.append({'offer_no': str(data[0]),
                             'subject_code': str(data[1]),
                                               'school_days':str(data[3]), 'department_code':str(data[7]), 
                                               'subject_title': data[2], 'faculty_id':str(data[8]),
                                               'school_time': str(data[4]), 'sem_id': str(data[5]), 
                                               'academic_year': str(data[6])})
                    check_subjectOfferings = SubjectOfferings.objects.all()
                    is_subjectOfferings = bool(check_subjectOfferings)
                    get_subject_code = Subject.objects.get(
                        subject_code=data[1])
                    get_department_code = Department.objects.get(
                        department_code=data[7])
                    get_faculty_id = Faculty.objects.get(
                        faculty_id=str(data[8]))

                    if is_subjectOfferings:
                        try:
                            check_if_exist = SubjectOfferings.objects.get(
                                offer_no=str(data[0]), sem_id=str(data[5]),academic_year = str(data[6]))
                            check_exist = bool(check_if_exist)
                        except Exception:
                            check_exist = False
                        if check_exist:
                            if sem_choice == check_if_exist.sem_id and sem_year == check_if_exist.academic_year:
                                check_if_exist.offer_no = str(data[0])
                                check_if_exist.subject_code = get_subject_code
                                check_if_exist.subject_title = data[2]
                                check_if_exist.school_time = str(data[4])
                                check_if_exist.sem_id = str(data[5])
                                check_if_exist.academic_year = str(data[6])
                                check_if_exist.faculty_id = get_faculty_id
                                check_if_exist.save()
                        else:
                            if sem_choice == str(data[5]) and sem_year == str(data[6]):
                                value = SubjectOfferings(
                                    offer_no=str(data[0]),
                                    subject_code=get_subject_code,
                                    subject_title=data[2],
                                    school_days=str(data[3]),
                                    school_time=str(data[4]),
                                    sem_id=str(data[5]),
                                    academic_year=str(data[6]),
                                    department_code=get_department_code,
                                    faculty_id=get_faculty_id
                                )
                                value.save()
                    else:
                        if sem_choice == str(data[5]) and sem_year == str(data[6]):
                            value = SubjectOfferings(
                                offer_no=str(data[0]),
                                subject_code=get_subject_code,
                                subject_title=data[2],
                                school_days=str(data[3]),
                                school_time=str(data[4]),
                                sem_id=str(data[5]),
                                academic_year=str(data[6]),
                                department_code=get_department_code,
                                faculty_id=get_faculty_id
                            )
                            value.save()

                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
        page = request.GET.get('page', 1)

        paginator = Paginator(new_subject_offerings_list, 10)
        try:
            subject_offerings = paginator.page(page)
        except PageNotAnInteger:
            subject_offerings = paginator.page(1)
        except EmptyPage:
            subject_offerings = paginator.page(paginator.num_pages)
    except Exception as e:
        subject_offerings=[]
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_subject_offerings.html", {'sem_choice':sem_choice,'sem_year':sem_year,"subject_offerings": subject_offerings, 'check_sem': check_sem, 'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def upload_student(request):
    global Active_Year
    global Active_Sem
    global new_student_list
    try:
        get_student = Student.objects.all()
        if request.method == 'POST':
            StudentResource()
            dataset = Dataset()
            new_student = request.FILES['myfile']
            imported_data = dataset.load(
                new_student.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_student)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 11):
                for data in imported_data:
                    try:
                        check_if_exist_school = School.objects.get(
                            school_name=str(data[4]))
                    except Exception:
                        value = School(school_name=str(data[4]))
                        value.save()
                    try:
                        check_if_exist_degree = DegreeProgram.objects.get(
                            program_code=str(data[6]))
                    except Exception:
                        value = DegreeProgram(program_code=str(data[6]))
                        value.save()
                    try:
                        get_department = Department.objects.get(
                            department_code=str(data[5]))
                    except Exception:
                        depa = Department(department_code=str(data[5]))
                        depa.save()
                    get_department = Department.objects.get(
                        department_code=str(data[5]))
                    get_school = School.objects.get(
                        school_name=str(data[4]))
                    get_degree = DegreeProgram.objects.get(
                        program_code=str(data[6]))
                    value = Student(
                        student_number=str(data[0]),
                        lastname=data[1],
                        firstname=data[2],
                        middlename=data[3],
                        school_name=get_school,
                        department_code=get_department,
                        program_code=get_degree,
                        academic_year=str(data[7]),
                        sem_id=data[8],
                        student_email=data[9],
                        role=data[10]
                    )
                    value.save()
                    new_student_list.append({'student_number': str(data[0]), 'lastname': data[1],
                                               'firstname': data[2], 'middlename': data[3], 'school_name': str(data[4]),
                                               'program_code':str(data[6]),'department_code':str(data[5]),
                                                'academic_year': str(data[7]),'sem_id':data[8],
                                                 'role': data[10],'student_email':data[9]})
                        
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
        page = request.GET.get('page', 1)

        paginator = Paginator(new_student_list, 10)
        try:
            student = paginator.page(page)
        except PageNotAnInteger:
            student = paginator.page(1)
        except EmptyPage:
            student = paginator.page(paginator.num_pages)
    except Exception as e:
        student=[]
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_student.html", {"student": student,'Active_Year':Active_Year,'Active_Sem':Active_Sem})


@login_required(login_url='login')
def upload_student_load(request):
    global Active_Year
    global Active_Sem
    global new_student_load_list
    try:
        get_student_load = Studentload.objects.all()
        if request.method == 'POST':
            StudentloadResource()
            dataset = Dataset()
            new_student_load = request.FILES['myfile']
            imported_data = dataset.load(
                new_student_load.read(), format='xlsx')
            wb_obj = openpyxl.load_workbook(new_student_load)
            sheet_obj = wb_obj.active
            col = sheet_obj.max_column
            row = sheet_obj.max_row

            if(col == 4):
                for data in imported_data:
                    new_student_load_list.append({'student_number': str(data[0]), 
                        'offer_no': str(data[1]), 'sem_id':str(data[2]),'academic_year': str(data[3])})
                    check_student_load = Studentload.objects.all()
                    is_student_load = bool(check_student_load)
                    get_student_number = Student.objects.get(
                        student_number=str(data[0]))
                    get_offer_no = SubjectOfferings.objects.get(
                        offer_no=str(data[1]), sem_id=str(data[2]))
                    if is_student_load:
                        try:
                            check_if_exist = Studentload.objects.get(
                                student_number=get_student_number, offer_no=get_offer_no)
                            check = bool(check_if_exist)
                        except Exception:
                            check = False
                        if check:
                            check_if_exist.student_number = get_student_number
                            check_if_exist.offer_no = get_offer_no
                            check_if_exist.sem_id = data[2]
                            check_if_exist.academic_year = str(data[3])
                            check_if_exist.save()
                        else:
                            value = Studentload(student_number=get_student_number,
                                                offer_no=get_offer_no,
                                                sem_id=str(data[2]),
                                                academic_year=str(data[3])
                                                )
                            value.save()
                    else:
                        value = Studentload(
                            student_number=get_student_number,
                            offer_no=get_offer_no,
                            sem_id=str(data[2]),
                            academic_year=str(data[3]))
                        value.save()
                        
                messages.info(request, 'Successfully Added')
            else:
                messages.info(request, 'Failed to Add the Data')
        page = request.GET.get('page', 1)

        paginator = Paginator(new_student_load_list, 10)
        try:
            student_load = paginator.page(page)
        except PageNotAnInteger:
            student_load = paginator.page(1)
        except EmptyPage:
            student_load = paginator.page(paginator.num_pages)
    except Exception as e:
        print('eeee',e)
        student_load = []
        messages.info(request, 'Please Choose File')
    return render(request, "admin/upload_student_load.html", {"student_load": student_load,'Active_Year':Active_Year,'Active_Sem':Active_Sem})
# admin


# director
@login_required(login_url='login')
def director_home_view(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    return render(request, "director/home.html", {"form": director_name})


@login_required(login_url='login')
def list_degree_program(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    degree_program = DegreeProgram.objects.all().order_by('program_code')
    return render(request, "director/list_degree_program.html", {"form": director_name, 'degree_program': degree_program})


@login_required(login_url='login')
def assign_counselor(request, code):
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    degree_program = DegreeProgram.objects.get(program_code=code)
    edit_form = AssignCounselorForm()
    if request.method == "POST":
        edit_form = AssignCounselorForm(request.POST)
        if edit_form.is_valid():
            new_faculty = edit_form['faculty'].value()
            get_counselor = Faculty.objects.get(faculty_id=new_faculty)
            edit = DegreeProgram.objects.get(program_code=code)
            edit.faculty_id = get_counselor
            edit.save()
            degree_program = DegreeProgram.objects.get(program_code=code)
    return render(request, "director/assign_counselor.html", {"form": director_name, 'degree_program': degree_program, 'edit_form': edit_form})


@login_required(login_url='login')
def per_degree_program(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    degree_program = DegreeProgram.objects.all().order_by('program_code')
    return render(request, "director/per_degree_program.html", {"form": director_name, 'degree_program': degree_program})


@login_required(login_url='login')
def view_stat_by_degree_program(request, degree):
    global filterStartDate
    global filterEndData
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    referrals = Referral.objects.filter(degree_program=degree)
    stat = referrals.count()
    filterDate = FilterDateForm()
    if request.method == "POST":
        filterDate = FilterDateForm(request.POST)
        if filterDate.is_valid():
            start = filterDate['pickedStartDate'].value()
            end = filterDate['pickedEndDate'].value()
            filterStartDate = start
            filterEndData = end
    degree = DegreeProgram.objects.get(program_code=degree)
    return render(request, "director/view_stat_by_degree_program.html", {'filterDate': filterDate, 'degree': degree, "referrals": referrals, "form": director_name, 'stat': stat})


@login_required(login_url='login')
def view_stat_by_degree_program_with_date(request, degree):
    global filterStartDate
    global filterEndData
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    filterDate = FilterDateForm()
    if request.method == "POST":
        filterDate = FilterDateForm(request.POST)
        if filterDate.is_valid():
            start = filterDate['pickedStartDate'].value()
            end = filterDate['pickedEndDate'].value()
            filterStartDate = start
            filterEndData = end
    degree = DegreeProgram.objects.get(program_code=degree)
    students = Referral.objects.all()
    referrals = []
    newStartDate = datetime.strptime(filterStartDate, "%Y-%m-%d").date()
    newEndDate = datetime.strptime(filterEndData, "%Y-%m-%d").date()
    for obj in students:
        if obj.degree_program == degree:
            if obj.date >= newStartDate and obj.date <= newEndDate:
                referrals.append(Referral(firstname=obj.firstname,
                                          lastname=obj.lastname, student_number=obj.student_number,
                                          degree_program=obj.degree_program, subject_referred=obj.subject_referred,
                                          reasons=obj.reasons, behavior_problem=obj.behavior_problem, date=obj.date,
                                          feedback=obj.feedback))
    stat = len(referrals)
    return render(request, "director/view_stat_by_degree_program_with_date.html", {'start': filterStartDate, 'end': filterEndData, 'filterDate': filterDate, 'degree': degree, "referrals": referrals, "form": director_name, 'stat': stat})


@login_required(login_url='login')
def per_counselor(request, *args, **kwargs):
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    counselor = Faculty.objects.filter(role='Counselor').order_by('lastname')
    return render(request, "director/per_counselor.html", {"form": director_name, 'counselor': counselor})


@login_required(login_url='login')
def view_stat_by_counselor(request, counselor_id):
    global filterStartDate
    global filterEndData
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    referrals = Referral.objects.filter(counselor_id=counselor_id)
    stat = referrals.count()
    counselor_fullname = Faculty.objects.get(
        faculty_id=counselor_id)
    filterDate = FilterDateForm()
    if request.method == "POST":
        filterDate = FilterDateForm(request.POST)
        if filterDate.is_valid():
            start = filterDate['pickedStartDate'].value()
            end = filterDate['pickedEndDate'].value()
            filterStartDate = start
            filterEndData = end
    return render(request, "director/view_stat_by_counselor.html", {'couns': counselor_fullname, 'filterDate': filterDate, 'counselor_id': counselor_id, "referrals": referrals, "form": director_name, 'stat': stat})


@login_required(login_url='login')
def view_stat_by_counselor_with_date(request, counselor_id):
    global filterStartDate
    global filterEndData
    user = request.session.get('username')
    director_name = Faculty.objects.get(faculty_id=user)
    filterDate = FilterDateForm()
    if request.method == "POST":
        filterDate = FilterDateForm(request.POST)
        if filterDate.is_valid():
            start = filterDate['pickedStartDate'].value()
            end = filterDate['pickedEndDate'].value()
            filterStartDate = start
            filterEndData = end
            filterDate = FilterDateForm()
    counselor_fullname = Faculty.objects.get(
        faculty_id=counselor_id)
    students = Referral.objects.all()
    detail = []
    newStartDate = datetime.strptime(filterStartDate, "%Y-%m-%d").date()
    newEndDate = datetime.strptime(filterEndData, "%Y-%m-%d").date()
    for obj in students:
        if obj.counselor_id == counselor_id:
            if obj.date >= newStartDate and obj.date <= newEndDate:
                detail.append(Referral(firstname=obj.firstname,
                                       lastname=obj.lastname, student_number=obj.student_number,
                                       degree_program=obj.degree_program, subject_referred=obj.subject_referred,
                                       reasons=obj.reasons, behavior_problem=obj.behavior_problem, date=obj.date,
                                       feedback=obj.feedback))
    stat = len(detail)
    return render(request, "director/view_stat_by_counselor_with_date.html", {'filterDate': filterDate, 'start': filterStartDate, 'end': filterEndData, 'counselor_id': counselor_id, 'couns': counselor_fullname, 'form': director_name, 'stat_details': detail, 'stat': stat})


# director


# counselor
@login_required(login_url='login')
def counselor_home_view(request, *args, **kwargs):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    return render(request, "counselor/home.html", {"counselorNotif": counselorNotif, "form": counselor_name})



@login_required(login_url='login')
def view_referred_students(request):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    qs = Referral.objects.filter(counselor_id=user, status='done')
    return render(request, "counselor/view_referred_students.html", {"counselorNotif": counselorNotif, "objects": qs, "form": counselor_name})


@login_required(login_url='login')
def view_pending_referred_students(request):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    qs = Referral.objects.filter(counselor_id=user, status='pending')
    return render(request, "counselor/view_pending_referred_students.html", {"counselorNotif": counselorNotif, "objects": qs, "form": counselor_name})


@login_required(login_url='login')
def detail_referred_student_counselor(request, id):
    print('hoy')
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    qs=[]
    get_referral = Referral.objects.get(id=id)
    for obj in get_referral.referral_id:
            get_details = ReferralDetails.objects.get(id=obj)
            qs.append(ReferralDetails(subject_referred=get_details.subject_referred,
                                            reasons=get_details.reasons, behavior_problem=get_details.behavior_problem,
                                            faculty_id=get_details.faculty_id))
    count = len(get_referral.referral_id)
    referral_id = get_referral.id
    return render(request, "counselor/detail_referred_student_counselor.html", {"counselorNotif": counselorNotif,'referral_id':referral_id, 'get_referral':get_referral, 'count': count, "detail": qs, 'get_referral':get_referral, "form": counselor_name})


@login_required(login_url='login')
def counselor_set_schedule(request, *args, **kwargs):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    offer = SetScheduleCounselorForm()
    if request.method == "POST":
        offer = SetScheduleCounselorForm(request.POST)
        if offer.is_valid():
            print('valid')
            pickedDateForm = offer['date'].value()
            start_timeForm = offer['start_time'].value()
            end_timeForm = offer['end_time'].value()

            date = datetime.strptime(pickedDateForm, "%Y-%m-%d").date()
            ClassesCounselor = []
            classes_of_counselor_checker = False
            timeStartNotAvailable = datetime.strptime(
                start_timeForm + ':00', '%H:%M:%S').time()
            timeEndNotAvailable = datetime.strptime(
                end_timeForm + ':00', '%H:%M:%S').time()
            day_name = date.strftime("%a")
            if timeStartNotAvailable >= timeEndNotAvailable:
                # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                messages.info(request, 'Conflict of Time')
            else:
                classes_of_counselor = SubjectOfferings.objects.filter(
                    faculty_id=user)
                classes_of_counselor_checker = bool(classes_of_counselor)

                if (classes_of_counselor_checker == True):
                    for object in classes_of_counselor:
                        check = bool(day_name[0].upper() in object.school_days)
                        if check == True:
                            ClassesCounselor.append(SubjectOfferings(offer_no=object.offer_no, school_days=object.school_days,
                                                                     school_time=object.school_time,
                                                                     subject_code=object.subject_code, sem_id=object.sem_id, academic_year=object.academic_year,))
                not_available_sched = SetScheduleCounselor.objects.filter(
                    faculty_id=user, date=date)
                not_available_sched_checker = bool(not_available_sched)

                checker = 0
                checker1 = 0
                checker2 = 0
                referral_by_datenotAvailable = Referral.objects.filter(counselor_id=user,
                                                                       date=date).order_by('start_time')

                referral_by_datenotAvailable_checker = bool(
                    referral_by_datenotAvailable)

                # testig part
                # referral_by_datenotAvailable_checker = False
                # not_available_sched_checker = False
                # classes_of_counselor_checker = False
                if(referral_by_datenotAvailable_checker == True and classes_of_counselor_checker == True and not_available_sched_checker == True):
                    print('1')
                    for object in ClassesCounselor:
                        get_time = object.school_time
                        classes_counselor_time = get_time.split(
                            '-')
                        classes_counselor_start_time = classes_counselor_time[0].upper(
                        ).replace(" ", "")
                        classes_counselor_end_time = classes_counselor_time[1].upper(
                        ).replace(" ", "")

                        # for start_time
                        if (classes_counselor_start_time[1]) == ':':
                            classes_counselor_start_time = "".join(
                                ('0', classes_counselor_start_time))
                        if classes_counselor_start_time[-2:] == "AM":
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = str(
                                    '00' + classes_counselor_start_time[2:-2])
                            else:
                                cc_start = classes_counselor_start_time[:-2]
                        else:
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = classes_counselor_start_time[:-2]
                            else:
                                cc_start = str(
                                    int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                        # for end_time
                        if classes_counselor_end_time[1] == ':':
                            classes_counselor_end_time = "".join(
                                ('0', classes_counselor_end_time))
                        if classes_counselor_end_time[-2:] == "AM":
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = str(
                                    '00' + classes_counselor_end_time[2:-2])
                            else:
                                cc_end = classes_counselor_end_time[:-2]
                        else:
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = classes_counselor_end_time[:-2]
                            else:
                                cc_end = str(
                                    int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                        cc_start_convert = datetime.strptime(
                            cc_start, '%H:%M').time()
                        cc_end_convert = datetime.strptime(
                            cc_end, '%H:%M').time()
                        # sud ni sha sa for object
                        if(cc_start_convert <= timeStartNotAvailable and cc_end_convert >= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_start_convert < cc_end_convert or cc_end_convert > timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):  # tupong ni sa for object
                        for object1 in referral_by_datenotAvailable:
                            if(object1.start_time <= timeStartNotAvailable and object1.end_time >= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.start_time < timeEndNotAvailable or object1.end_time > timeStartNotAvailable and object1.end_time <= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.end_time <= timeEndNotAvailable):
                                checker1 = 1
                        if(checker1 != 1):  # tupong sa for object1
                            for object2 in not_available_sched:
                                if(object2.start_time <= timeStartNotAvailable and object2.end_time >= timeEndNotAvailable or object2.start_time >= timeStartNotAvailable and object2.start_time < timeEndNotAvailable or object2.end_time > timeStartNotAvailable and object2.end_time <= timeEndNotAvailable or object2.start_time >= timeStartNotAvailable and object2.end_time <= timeEndNotAvailable):
                                    checker2 = 1
                            if(checker2 != 1):
                                # -- save the id date and time in the database
                                offer.save()
                                newData = SetScheduleCounselor.objects.last()
                                newData.faculty_id = user
                                newData.choice = 'Not Available'
                                newData.save()
                                # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                                messages.info(request, 'Success')
                            else:  # else checker2
                                messages.info(
                                    request, 'Not Available Time')
                        else:  # else checker1
                            messages.info(
                                request, 'Not Available Time')
                    else:  # else checker
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == True and classes_of_counselor_checker == True and not_available_sched_checker == False):
                    print('2')
                    for object in ClassesCounselor:
                        get_time = object.school_time
                        classes_counselor_time = get_time.split(
                            '-')
                        classes_counselor_start_time = classes_counselor_time[0].upper(
                        ).replace(" ", "")
                        classes_counselor_end_time = classes_counselor_time[1].upper(
                        ).replace(" ", "")

                        # for start_time
                        if (classes_counselor_start_time[1]) == ':':
                            classes_counselor_start_time = "".join(
                                ('0', classes_counselor_start_time))
                        if classes_counselor_start_time[-2:] == "AM":
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = str(
                                    '00' + classes_counselor_start_time[2:-2])
                            else:
                                cc_start = classes_counselor_start_time[:-2]
                        else:
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = classes_counselor_start_time[:-2]
                            else:
                                cc_start = str(
                                    int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                        # for end_time
                        if classes_counselor_end_time[1] == ':':
                            classes_counselor_end_time = "".join(
                                ('0', classes_counselor_end_time))
                        if classes_counselor_end_time[-2:] == "AM":
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = str(
                                    '00' + classes_counselor_end_time[2:-2])
                            else:
                                cc_end = classes_counselor_end_time[:-2]
                        else:
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = classes_counselor_end_time[:-2]
                            else:
                                cc_end = str(
                                    int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                        cc_start_convert = datetime.strptime(
                            cc_start, '%H:%M').time()
                        cc_end_convert = datetime.strptime(
                            cc_end, '%H:%M').time()
                        # sud ni sha sa for object
                        if(cc_start_convert <= timeStartNotAvailable and cc_end_convert >= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_start_convert < cc_end_convert or cc_end_convert > timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        for object1 in referral_by_datenotAvailable:
                            if(object1.start_time <= timeStartNotAvailable and object1.end_time >= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.start_time < timeEndNotAvailable or object1.end_time > timeStartNotAvailable and object1.end_time <= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.end_time <= timeEndNotAvailable):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.faculty_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'Success')
                        else:
                            messages.info(
                                request, 'Not Available Time')
                    else:
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == False and classes_of_counselor_checker == True and not_available_sched_checker == True):
                    print('3')
                    for object in ClassesCounselor:
                        get_time = object.school_time
                        classes_counselor_time = get_time.split(
                            '-')
                        classes_counselor_start_time = classes_counselor_time[0].upper(
                        ).replace(" ", "")
                        classes_counselor_end_time = classes_counselor_time[1].upper(
                        ).replace(" ", "")

                        # for start_time
                        if (classes_counselor_start_time[1]) == ':':
                            classes_counselor_start_time = "".join(
                                ('0', classes_counselor_start_time))
                        if classes_counselor_start_time[-2:] == "AM":
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = str(
                                    '00' + classes_counselor_start_time[2:-2])
                            else:
                                cc_start = classes_counselor_start_time[:-2]
                        else:
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = classes_counselor_start_time[:-2]
                            else:
                                cc_start = str(
                                    int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                        # for end_time
                        if classes_counselor_end_time[1] == ':':
                            classes_counselor_end_time = "".join(
                                ('0', classes_counselor_end_time))
                        if classes_counselor_end_time[-2:] == "AM":
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = str(
                                    '00' + classes_counselor_end_time[2:-2])
                            else:
                                cc_end = classes_counselor_end_time[:-2]
                        else:
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = classes_counselor_end_time[:-2]
                            else:
                                cc_end = str(
                                    int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                        cc_start_convert = datetime.strptime(
                            cc_start, '%H:%M').time()
                        cc_end_convert = datetime.strptime(
                            cc_end, '%H:%M').time()
                        # sud ni sha sa for object
                        if(cc_start_convert <= timeStartNotAvailable and cc_end_convert >= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_start_convert < cc_end_convert or cc_end_convert > timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        for object1 in not_available_sched:
                            if(object1.start_time <= timeStartNotAvailable and object1.end_time >= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.start_time < timeEndNotAvailable or object1.end_time > timeStartNotAvailable and object1.end_time <= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.end_time <= timeEndNotAvailable):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.faculty_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'Success')
                        else:
                            messages.info(
                                request, 'Not Available Time')
                    else:
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == True and classes_of_counselor_checker == False and not_available_sched_checker == True):
                    print('4')
                    for object in referral_by_datenotAvailable:
                        # sud ni sha sa for object
                        if(object.start_time <= timeStartNotAvailable and object.end_time >= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.start_time < timeEndNotAvailable or object.end_time > timeStartNotAvailable and object.end_time <= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.end_time <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        for object1 in not_available_sched:
                            if(object1.start_time <= timeStartNotAvailable and object1.end_time >= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.start_time < timeEndNotAvailable or object1.end_time > timeStartNotAvailable and object1.end_time <= timeEndNotAvailable or object1.start_time >= timeStartNotAvailable and object1.end_time <= timeEndNotAvailable):
                                checker1 = 1
                        if(checker1 != 1):
                            offer.save()
                            newData = SetScheduleCounselor.objects.last()
                            newData.faculty_id = user
                            newData.choice = 'Not Available'
                            newData.save()
                            # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                            messages.info(request, 'Success')
                        else:
                            messages.info(
                                request, 'Not Available Time')
                    else:
                        messages.info(
                            request, 'Not Available Time')
                if(referral_by_datenotAvailable_checker == False and classes_of_counselor_checker == True and not_available_sched_checker == False):
                    print('5')
                    for object in ClassesCounselor:
                        get_time = object.school_time
                        classes_counselor_time = get_time.split(
                            '-')
                        classes_counselor_start_time = classes_counselor_time[0].upper(
                        ).replace(" ", "")
                        classes_counselor_end_time = classes_counselor_time[1].upper(
                        ).replace(" ", "")

                        # for start_time
                        if (classes_counselor_start_time[1]) == ':':
                            classes_counselor_start_time = "".join(
                                ('0', classes_counselor_start_time))
                        if classes_counselor_start_time[-2:] == "AM":
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = str(
                                    '00' + classes_counselor_start_time[2:-2])
                            else:
                                cc_start = classes_counselor_start_time[:-2]
                        else:
                            if classes_counselor_start_time[:2] == '12':
                                cc_start = classes_counselor_start_time[:-2]
                            else:
                                cc_start = str(
                                    int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                        # for end_time
                        if classes_counselor_end_time[1] == ':':
                            classes_counselor_end_time = "".join(
                                ('0', classes_counselor_end_time))
                        if classes_counselor_end_time[-2:] == "AM":
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = str(
                                    '00' + classes_counselor_end_time[2:-2])
                            else:
                                cc_end = classes_counselor_end_time[:-2]
                        else:
                            if classes_counselor_end_time[:2] == '12':
                                cc_end = classes_counselor_end_time[:-2]
                            else:
                                cc_end = str(
                                    int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                        cc_start_convert = datetime.strptime(
                            cc_start, '%H:%M').time()
                        cc_end_convert = datetime.strptime(
                            cc_end, '%H:%M').time()
                        # sud ni sha sa for object
                        if(cc_start_convert <= timeStartNotAvailable and cc_end_convert >= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_start_convert < cc_end_convert or cc_end_convert > timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable or cc_start_convert >= timeStartNotAvailable and cc_end_convert <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.faculty_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'Success')
                    else:
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == True and classes_of_counselor_checker == False and not_available_sched_checker == False):
                    print('6')
                    for object in referral_by_datenotAvailable:
                        # sud ni sha sa for object
                        if(object.start_time <= timeStartNotAvailable and object.end_time >= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.start_time < timeEndNotAvailable or object.end_time > timeStartNotAvailable and object.end_time <= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.end_time <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.faculty_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'Success')
                    else:
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == False and classes_of_counselor_checker == False and not_available_sched_checker == True):
                    print('7')
                    for object in not_available_sched:
                        # sud ni sha sa for object
                        if(object.start_time <= timeStartNotAvailable and object.end_time >= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.start_time < timeEndNotAvailable or object.end_time > timeStartNotAvailable and object.end_time <= timeEndNotAvailable or object.start_time >= timeStartNotAvailable and object.end_time <= timeEndNotAvailable):
                            checker = 1
                    if(checker != 1):
                        offer.save()
                        newData = SetScheduleCounselor.objects.last()
                        newData.faculty_id = user
                        newData.choice = 'Not Available'
                        newData.save()
                        # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                        messages.info(request, 'Success')
                    else:
                        messages.info(
                            request, 'Not Available Time')

                if(referral_by_datenotAvailable_checker == False and classes_of_counselor_checker == False and not_available_sched_checker == False):
                    print('9')
                    offer.save()
                    newData = SetScheduleCounselor.objects.last()
                    newData.faculty_id = user
                    newData.choice = 'Not Available'
                    newData.save()
                    # offer = SetScheduleCounselorForm(request.POST, initial={'employee_id': user, 'choice': 'Not Available'})
                    messages.info(request, 'Success')
    return render(request, "counselor/counselor_set_schedule.html", {"offer": offer, "counselorNotif": counselorNotif, "form": counselor_name})


@login_required
def counselor_notifications(request):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = Notification.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == Notification.AUTOMATIC_REFERRAL:
            return render(request, "counselor/counselor.html", {})
        elif notification.notification_type == Notification.MANUAL_REFERRAL:
            return render(request, "counselor/counselor.html", {})

    counselorNotif = Notification.objects.filter(
        to_user=user).order_by('created_at')
    return render(request, 'counselor/notification.html', {"notifications": counselorNotif, "form": counselor_name})


@login_required(login_url='login')
def counselor_notification_detail(request, id):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    notification = Notification.objects.get(id=id)
    notification.is_read_counselor = True
    notification.save()
    detail = []
    get_referral = Referral.objects.get(id=id)
    for obj in get_referral.referral_id:
            get_details = ReferralDetails.objects.get(id=obj)
            detail.append(ReferralDetails(subject_referred=get_details.subject_referred,
                                            reasons=get_details.reasons, behavior_problem=get_details.behavior_problem,
                                            faculty_id=get_details.faculty_id))
    
    referral_id = get_referral.id
    return render(request, "counselor/counselor_notification_detail.html", {"counselorNotif": counselorNotif, 'get_referral':get_referral, "detail": detail, "referral_id": referral_id, "form": counselor_name})


@login_required(login_url='login')
def counselor_feedback_student(request, id):
    print('hakdog one kung si teacher nag refer')
    global feedback_id
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    info = Notification.objects.get(id=id)
    student = Referral.objects.get(id=id)
    preparedby = Faculty.objects.get(faculty_id=student.counselor_id)
    form1 = CounselorFeedbackForm()
    teachers=[]
    if request.method == "POST":
        form1 = CounselorFeedbackForm(request.POST)
        if form1.is_valid():
            form1.save()
            feedback = form1['remarks'].value()
            t = Referral.objects.get(id=id)
            t.status = "done"
            t.feedback = feedback
            t.save()
            for obj in student.referral_id:
                    get_detail = ReferralDetails.objects.get(id=obj)
                    if get_detail.faculty_id is not None:
                        try:
                            check = NotificationFeedback.objects.get(extra_id=id)
                            check = True
                        except:
                            check = False
                        if check==False:
                            teachers.append(get_detail.faculty_id_id)
            for obj in teachers:
                create_feedback(obj,'feedback_teacher', user, id, feedback_id)
            
            messages.info(request, 'Success!')
            form1 = CounselorFeedbackForm()
            return render(request, "counselor/counselor_feedback_student.html", {"counselorNotif": counselorNotif, "info": info,  "info2": preparedby,   "student": student, "object": form1, "form": counselor_name})

    return render(request, "counselor/counselor_feedback_student.html", {"counselorNotif": counselorNotif, "info": info, "info2": preparedby, "student": student, "object": form1, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_feedback(request):
    global Search
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    student = Referral.objects.filter(counselor_id=user, status='done')
    date = CalendarForm()
    if request.method == "POST":
        date = CalendarForm(request.POST)
        if date.is_valid():
            date_choice = date['pickedDate'].value()
            date.save()
            return redirect('counselor_view_feedback_with_date', date=date_choice)
    return render(request, 'counselor/view_feedback.html', {"date": date, "counselorNotif": counselorNotif, "student": student, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_feedback_with_date(request, date):
    global Search
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    student = Referral.objects.filter(
        counselor_id=user, status='done', date=date)
    date = CalendarForm()
    if request.method == "POST":
        date = CalendarForm(request.POST)
        if date.is_valid():
            date_choice = date['pickedDate'].value()
            date.save()
            return redirect('counselor_view_feedback_with_date', date=date_choice)
    return render(request, 'counselor/view_feedback_with_date.html', {"date": date, "counselorNotif": counselorNotif, "student": student, "form": counselor_name})


@login_required(login_url='login')
def detail_referred_student_with_feedback(request, id):
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    qs = Referral.objects.get(id=id)

    return render(request, "counselor/detail_referred_student_with_feedback.html", {"counselorNotif": counselorNotif, "object": qs, "form": counselor_name})


@login_required(login_url='login')
def counselor_view_schedule(request, *args, **kwargs):
    global Active_Sem
    global Active_Year
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    global count
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    count = 0
    today = date.today()
    now = dt.datetime.now()
    day_name = now.strftime("%a")
    classes_of_counselor = []
    classes_of_counselor_checl = False

    ScheduledReferralbyDayCheck = False

    classes_of_counselor_list = SubjectOfferings.objects.filter(
        faculty_id=user,sem_id = Active_Sem , academic_year=Active_Year)
    classes_of_counselor_list_checker = bool(classes_of_counselor_list)

    if(classes_of_counselor_list_checker == True):
        for object in classes_of_counselor_list:
            if day_name == 'Thu':
                check = bool(
                    day_name[:-1].upper() in object.school_days)
            else:
                check = bool(
                    day_name[0].upper() in object.school_days)
            if(check == True):
                classes_of_counselor.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                                                             subject_title=object.subject_title, school_days=object.school_days,
                                                             school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                                                             department_code=object.department_code, faculty_id=object.faculty_id))
    else:
        classes_of_counselor_list_checker = False

    classes_counselor_checker = bool(classes_of_counselor)

    referral_list_byday = Referral.objects.filter(
        counselor_id=user, date=today).order_by('start_time')
    referral_list_byday_checker = bool(referral_list_byday)

    not_available_sched = SetScheduleCounselor.objects.filter(
        faculty_id=user, date=today)
    not_available_sched_checker = bool(not_available_sched)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    # one = []
    schedule_for_today = []

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('7:00:00', '%H:%M:%S').time()
    end = datetime.strptime('18:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    
    initialtime = 7
    id = 0
    for x in range(12):
            id +=1
            newTime = str(initialtime)+':00:00'
            time1 = datetime.strptime(newTime, '%H:%M:%S').time()
            newTime = str(initialtime)+':30:00'
            time2 = datetime.strptime(newTime, '%H:%M:%S').time()
            initialtime += 1
            time = NewTime(time_id=id, time1=time1, time2=time2)
            time.save()
            flag=1
            while(flag==1):
                id +=1
                newTime = str(initialtime)+':00:00'
                next_time = datetime.strptime(newTime, '%H:%M:%S').time()
                time = NewTime(time_id=id, time1=time2, time2=next_time)
                time.save()
                flag=0
    alltime = NewTime.objects.all()
    if(classes_counselor_checker == True and referral_list_byday_checker == True and not_available_sched_checker == True):
        print('1')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                get_time = object.school_time
                classes_counselor_time = get_time.split('-')
                classes_counselor_start_time = classes_counselor_time[0].upper().replace(
                    " ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper().replace(
                    " ", "")
                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]

                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                # sud ni sha sa for object same sa all
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

                for object2 in referral_list_byday:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Counseling'
                        schedule_for_today.append(Referral(id=object2.id,student_number=object2.student_number, firstname=object2.firstname,
                                                           lastname=object2.lastname, middlename=object2.middlename,
                                                           degree_program=object2.degree_program, 
                                                           counselor_id=object2.counselor_id, 
                                                           status = object2.status,
                                                           start_time=object2.start_time, end_time=object2.end_time, date=object2.date,
                                                           choice=choice))
                for object3 in not_available_sched:
                    if(timeArray[x] == object3.start_time):
                        choice = 'Not Available'
                        schedule_for_today.append(SetScheduleCounselor(faculty_id=object3.faculty_id, date=object3.date, start_time=object3.start_time,
                                                                       end_time=object3.end_time, choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == True and not_available_sched_checker == False):
        print('2')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))
                for object2 in referral_list_byday:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Counseling'
                        schedule_for_today.append(Referral(id=object2.id,student_number=object2.student_number, firstname=object2.firstname,
                                                           lastname=object2.lastname, middlename=object2.middlename,
                                                           degree_program=object2.degree_program,counselor_id=object2.counselor_id, status = object2.status,
                                                           start_time=object2.start_time, end_time=object2.end_time, date=object2.date, 
                                                           choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('3')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

                for object2 in not_available_sched:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Not Available'
                        schedule_for_today.append(SetScheduleCounselor(faculty_id=object2.faculty_id, date=object2.date, start_time=object2.start_time,
                                                                       end_time=object2.end_time, choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == True and not_available_sched_checker == True):
        print('4')
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    schedule_for_today.append(Referral(id=object.id,student_number=object.student_number, firstname=object.firstname,
                                                           lastname=object.lastname, middlename=object.middlename,
                                                           degree_program=object.degree_program,counselor_id=object.counselor_id, status = object.status,
                                                           start_time=object.start_time, end_time=object.end_time, date=object.date, 
                                                           choice=choice))

            for object2 in not_available_sched:
                if(timeArray[x] == object2.start_time):
                    choice = 'Not Available'
                    schedule_for_today.append(SetScheduleCounselor(faculty_id=object2.faculty_id, date=object2.date, start_time=object2.start_time,
                                                                   end_time=object2.end_time, choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == False and not_available_sched_checker == False):
        print('5')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

    elif(classes_counselor_checker == False and referral_list_byday_checker == True and not_available_sched_checker == False):
        print('6')
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    schedule_for_today.append(Referral(id=object.id,student_number=object.student_number, firstname=object.firstname,
                                                           lastname=object.lastname, middlename=object.middlename,
                                                           degree_program=object.degree_program,counselor_id=object.counselor_id, status = object.status,
                                                           start_time=object.start_time, end_time=object.end_time, date=object.date, 
                                                           choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('7')
        for x in range(len(timeArray)):
            for object in not_available_sched:
                if(timeArray[x] == object.start_time):
                    choice = 'Not Available'
                    schedule_for_today.append(SetScheduleCounselor(faculty_id=object.faculty_id, date=object.date, start_time=object.start_time,
                                                                   end_time=object.end_time, choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('no data to show')

    counselor_name = Faculty.objects.get(faculty_id=user)

    return render(request, "counselor/counselor_view_schedule.html", {"offer": offer, "counselorNotif": counselorNotif, "today": today, "day_name": day_name, "schedForToday": schedule_for_today, "time": alltime, "form": counselor_name})


@login_required(login_url='login')
def another_counselor_view_schedule(request, *args, **kwargs):
    global Active_Sem
    global Active_Year
    user = request.session.get('username')
    counselor_name = Faculty.objects.get(faculty_id=user)
    notif = Notification.objects.filter(to_user=user, is_read_counselor=False)
    counselorNotif = len(notif)
    global count
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    count = 0
    newDate = Calendar.objects.last()
    today = newDate.pickedDate
    day_name = today.strftime("%a")
    classes_of_counselor = []

    classes_of_counselor_list = SubjectOfferings.objects.filter(
        faculty_id=user,sem_id = Active_Sem , academic_year=Active_Year)
    classes_of_counselor_list_checker = bool(classes_of_counselor_list)

    print('classes_of_counselor_list_checker',classes_of_counselor_list_checker)
    if(classes_of_counselor_list_checker == True):
        for object in classes_of_counselor_list:
            if day_name == 'Thu':
                check = bool(
                    day_name[:-1].upper() in object.school_days)
            else:
                check = bool(
                    day_name[0].upper() in object.school_days)
            if(check == True):
                classes_of_counselor.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                                                             subject_title=object.subject_title, school_days=object.school_days,
                                                             school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                                                             department_code=object.department_code, faculty_id=object.faculty_id))
    else:
        classes_of_counselor_list_checker = False

    classes_counselor_checker = bool(classes_of_counselor)

    referral_list_byday = Referral.objects.filter(
        counselor_id=user, date=today).order_by('start_time')
    referral_list_byday_checker = bool(referral_list_byday)

    not_available_sched = SetScheduleCounselor.objects.filter(
        faculty_id=user, date=today)
    not_available_sched_checker = bool(not_available_sched)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    # one = []
    schedule_for_today = []

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('7:00:00', '%H:%M:%S').time()
    end = datetime.strptime('18:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    
    initialtime = 7
    id = 0
    for x in range(12):
            id +=1
            newTime = str(initialtime)+':00:00'
            time1 = datetime.strptime(newTime, '%H:%M:%S').time()
            newTime = str(initialtime)+':30:00'
            time2 = datetime.strptime(newTime, '%H:%M:%S').time()
            initialtime += 1
            time = NewTime(time_id=id, time1=time1, time2=time2)
            time.save()
            flag=1
            while(flag==1):
                id +=1
                newTime = str(initialtime)+':00:00'
                next_time = datetime.strptime(newTime, '%H:%M:%S').time()
                time = NewTime(time_id=id, time1=time2, time2=next_time)
                time.save()
                flag=0
    alltime = NewTime.objects.all()
    if(classes_counselor_checker == True and referral_list_byday_checker == True and not_available_sched_checker == True):
        print('1')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                get_time = object.school_time
                classes_counselor_time = get_time.split('-')
                classes_counselor_start_time = classes_counselor_time[0].upper().replace(
                    " ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper().replace(
                    " ", "")
                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]

                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                # sud ni sha sa for object same sa all
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

                for object2 in referral_list_byday:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Counseling'
                        schedule_for_today.append(Referral(id=object2.id,student_number=object2.student_number, firstname=object2.firstname,
                                                           lastname=object2.lastname, middlename=object2.middlename,
                                                           degree_program=object2.degree_program, 
                                                           counselor_id=object2.counselor_id, status = object2.status,
                                                           start_time=object2.start_time, end_time=object2.end_time, date=object2.date,
                                                           choice=choice))
                for object3 in not_available_sched:
                    if(timeArray[x] == object3.start_time):
                        choice = 'Not Available'
                        schedule_for_today.append(SetScheduleCounselor(faculty_id=object3.faculty_id, date=object3.date, start_time=object3.start_time,
                                                                       end_time=object3.end_time, choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == True and not_available_sched_checker == False):
        print('2')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))
                for object2 in referral_list_byday:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Counseling'
                        schedule_for_today.append(Referral(id=object2.id,student_number=object2.student_number, firstname=object2.firstname,
                                                           lastname=object2.lastname, middlename=object2.middlename,
                                                           degree_program=object2.degree_program,counselor_id=object2.counselor_id, status = object2.status,
                                                           start_time=object2.start_time, end_time=object2.end_time, date=object2.date, 
                                                           choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('3')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

                for object2 in not_available_sched:
                    if(timeArray[x] == object2.start_time):
                        choice = 'Not Available'
                        schedule_for_today.append(SetScheduleCounselor(faculty_id=object2.faculty_id, date=object2.date, start_time=object2.start_time,
                                                                       end_time=object2.end_time, choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == True and not_available_sched_checker == True):
        print('4')
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    schedule_for_today.append(Referral(id=object.id,student_number=object.student_number, firstname=object.firstname,
                                                           lastname=object.lastname, middlename=object.middlename,
                                                           degree_program=object.degree_program,counselor_id=object.counselor_id, status = object.status,
                                                           start_time=object.start_time, end_time=object.end_time, date=object.date, 
                                                           choice=choice))

            for object2 in not_available_sched:
                if(timeArray[x] == object2.start_time):
                    choice = 'Not Available'
                    schedule_for_today.append(SetScheduleCounselor(faculty_id=object2.faculty_id, date=object2.date, start_time=object2.start_time,
                                                                   end_time=object2.end_time, choice=choice))

    elif(classes_counselor_checker == True and referral_list_byday_checker == False and not_available_sched_checker == False):
        print('5')
        for x in range(len(timeArray)):
            for object in classes_of_counselor:
                # kani sha tupong ni sha sa if(timeArray[x] == cc_start_convert ): sa itupong sab ang mag katupong ani basta sunda rani nga indention pareha rani sa referral
                get_time = object.school_time
                classes_counselor_time = get_time.split(
                    '-')
                classes_counselor_start_time = classes_counselor_time[0].upper(
                ).replace(" ", "")
                classes_counselor_end_time = classes_counselor_time[1].upper(
                ).replace(" ", "")

                # for start_time
                if (classes_counselor_start_time[1]) == ':':
                    classes_counselor_start_time = "".join(
                        ('0', classes_counselor_start_time))
                if classes_counselor_start_time[-2:] == "AM":
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = str(
                            '00' + classes_counselor_start_time[2:-2])
                    else:
                        cc_start = classes_counselor_start_time[:-2]
                else:
                    if classes_counselor_start_time[:2] == '12':
                        cc_start = classes_counselor_start_time[:-2]
                    else:
                        cc_start = str(
                            int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                # for end_time
                if classes_counselor_end_time[1] == ':':
                    classes_counselor_end_time = "".join(
                        ('0', classes_counselor_end_time))
                if classes_counselor_end_time[-2:] == "AM":
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = str(
                            '00' + classes_counselor_end_time[2:-2])
                    else:
                        cc_end = classes_counselor_end_time[:-2]
                else:
                    if classes_counselor_end_time[:2] == '12':
                        cc_end = classes_counselor_end_time[:-2]
                    else:
                        cc_end = str(
                            int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                cc_start_convert = datetime.strptime(
                    cc_start, '%H:%M').time()
                cc_end_convert = datetime.strptime(
                    cc_end, '%H:%M').time()
                if(timeArray[x] == cc_start_convert):
                    choice = 'Class'
                    schedule_for_today.append({'offer_no': object.offer_no, 'subject_code': object.subject_code,
                                               'school_days': object.school_days, 'department_code': object.department_code, 'subject_title': object.subject_title,
                                               'start_time': cc_start_convert, 'end_time': cc_end_convert, 'choice': 'Class'})
                    # schedule_for_today.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                    #                                            subject_title=object.subject_title, school_days=object.school_days,
                    #                                            school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                    #                                            department_code=object.department_code, faculty_id=object.faculty_id))

    elif(classes_counselor_checker == False and referral_list_byday_checker == True and not_available_sched_checker == False):
        print('6')
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    schedule_for_today.append(Referral(id=object.id,student_number=object.student_number, firstname=object.firstname,
                                                           lastname=object.lastname, middlename=object.middlename,
                                                           degree_program=object.degree_program,counselor_id=object.counselor_id, status = object.status,
                                                           start_time=object.start_time, end_time=object.end_time, date=object.date, 
                                                           choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('7')
        for x in range(len(timeArray)):
            for object in not_available_sched:
                if(timeArray[x] == object.start_time):
                    choice = 'Not Available'
                    schedule_for_today.append(SetScheduleCounselor(faculty_id=object.faculty_id, date=object.date, start_time=object.start_time,
                                                                   end_time=object.end_time, choice=choice))

    elif(classes_counselor_checker == False and referral_list_byday_checker == False and not_available_sched_checker == True):
        print('no data to show')

    counselor_name = Faculty.objects.get(faculty_id=user)
    return render(request, "counselor/counselor_view_schedule.html", {"offer": offer, "counselorNotif": counselorNotif, "today": today, "day_name": day_name, "schedForToday": schedule_for_today, "time": alltime, "form": counselor_name})


# counselor

# teacher


@login_required(login_url='login')
def teacher_home_view(request, *args, **kwargs):
    global Active_Sem
    global Active_Year
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    subjects = SubjectOfferings.objects.filter(
        faculty_id=teacher_name.faculty_id, sem_id = Active_Sem, academic_year = Active_Year).order_by('offer_no')
    return render(request, "teacher/home.html",  {'teacherNotif': teacherNotif, "form": teacher_name, 'subjects': subjects})


@login_required(login_url='login')
def student_list_enrolled(request, offer_no):
    global Active_Sem
    global Active_Year
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student_enroll', search=search_choice, offer_no=offer_no)
    get_offerings = SubjectOfferings.objects.get(offer_no=offer_no, sem_id = Active_Sem, academic_year=Active_Year)
    studentload = Studentload.objects.filter(offer_no=get_offerings)
    students = Student.objects.all().order_by('lastname')
    student_list = []
    for obj in students:
        for obj1 in studentload:
            if obj.student_number == obj1.student_number.student_number:
                student_list.append(Student(student_number=obj.student_number,
                                    lastname=obj.lastname, firstname=obj.firstname,
                                    middlename=obj.middlename, program_code=obj.program_code))
    page = request.GET.get('page', 1)

    paginator = Paginator(student_list, 10)
    try:
        list = paginator.page(page)
    except PageNotAnInteger:
        list = paginator.page(1)
    except EmptyPage:
        list = paginator.page(paginator.num_pages)
    
    return render(request, "teacher/student_list_enrolled.html",  {'teacherNotif': teacherNotif, 'search_form':search_form,"form": teacher_name, 'list': list, 'offer_no': offer_no})


@login_required(login_url='login')
def search_student_enroll(request, search,offer_no):
    global Active_Sem
    global Active_Year
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    search_form = SearchForm()
    if request.method == "POST":
        search_form = SearchForm(request.POST)
        if search_form.is_valid():
            search_choice = search_form['search'].value()
            return redirect('search_student_enroll', search=search_choice, offer_no=offer_no)
    get_offerings = SubjectOfferings.objects.get(offer_no=offer_no, sem_id = Active_Sem, academic_year=Active_Year)
    studentload = Studentload.objects.filter(offer_no=get_offerings)
    students = Student.objects.all().order_by('lastname')
    student_list = []
    new_student_list = []
    for obj in students:
        for obj1 in studentload:
            if obj.student_number == obj1.student_number.student_number:
                student_list.append(Student(student_number=obj.student_number,
                                    lastname=obj.lastname, firstname=obj.firstname,
                                    middlename=obj.middlename, program_code=obj.program_code))
    for check in student_list:
        if search.lower() in check.lastname.lower():
            new_student_list.append(Student(student_number=check.student_number,
                                                lastname=check.lastname, firstname=check.firstname,
                                                middlename=check.middlename, program_code=check.program_code))
    page = request.GET.get('page', 1)

    paginator = Paginator(new_student_list, 10)
    try:
        list = paginator.page(page)
    except PageNotAnInteger:
        list = paginator.page(1)
    except EmptyPage:
        list = paginator.page(paginator.num_pages)
    
    return render(request, "teacher/student_list_enrolled.html",  {'teacherNotif': teacherNotif, 'search_form':search_form,"form": teacher_name, 'list': list, 'offer_no': offer_no})



















@login_required(login_url='login')
def teacher_view_referred_students(request, status):
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    referrals = Referral.objects.all()
    referral_details = ReferralDetails.objects.all()
    qs = []
    id_list=[]
    if status == 'all' or status == '--':
        for obj in referrals:
            for check in obj.referral_id:
                print('check',check)
                get_details = ReferralDetails.objects.get(id = check)
                if get_details.faculty_id is not None:
                    print('huuuuyyy',get_details.faculty_id,get_details.faculty_id.faculty_id,obj.id)
                    if user == get_details.faculty_id.faculty_id:
                        id_list.append(obj.id)
        print(id_list)
        for flag in id_list:
                referral = Referral.objects.get(id=flag)
                qs.append(Referral(id=flag,firstname=referral.firstname,
                                                                lastname=referral.lastname, student_number=referral.student_number,
                                                                degree_program=referral.degree_program,
                                                                counselor_id=referral.counselor_id,
                                                                start_time=referral.start_time, end_time=referral.end_time,
                                                                date=referral.date,
                                                                status=referral.status))
    elif (status == 'pending'):
        for obj in referrals:
            if obj.status == 'pending':
                for check in obj.referral_id:
                    get_details = ReferralDetails.objects.get(id = check)
                    if get_details.faculty_id is not None:
                        if user == get_details.faculty_id.faculty_id:
                            id_list.append(obj.id)
                for flag in id_list:
                    referral = Referral.objects.get(id=flag)
                    qs.append(Referral(id=flag,firstname=referral.firstname,
                                                                    lastname=referral.lastname, student_number=referral.student_number,
                                                                    degree_program=referral.degree_program,
                                                                    counselor_id=referral.counselor_id,
                                                                    start_time=referral.start_time, end_time=referral.end_time,
                                                                    date=referral.date,
                                                                    status=referral.status))
        
    elif (status == 'done'):
        for obj in referrals:
            if obj.status == 'done':
                for check in obj.referral_id:
                    get_details = ReferralDetails.objects.get(id = check)
                    if get_details.faculty_id is not None:
                        if user == get_details.faculty_id.faculty_id:
                            id_list.append(obj.id)
                for flag in id_list:
                    referral = Referral.objects.get(id=flag)
                    qs.append(Referral(id=flag,firstname=referral.firstname,
                                                                    lastname=referral.lastname, student_number=referral.student_number,
                                                                    degree_program=referral.degree_program,
                                                                    counselor_id=referral.counselor_id,
                                                                    start_time=referral.start_time, end_time=referral.end_time,
                                                                    date=referral.date,
                                                                    status=referral.status))           
    # if status == 'all' or status == '--':
    #     print('all')
    #     for obj in referrals:
    #         for check in obj.referral_id:
    #             for obj1 in check:
    #                 print('obj',obj)
    #                 get_details = ReferralDetails.objects.get(id = obj1)
    #                 if get_details.faculty_id is not None:
    #                     if user == get_details.faculty_id.faculty_id:
    #                         flag = obj.id
    #     referral = Referral.objects.get(id=flag)
    #     qs.append(Referral(id=flag,firstname=referral.firstname,
    #                                                     lastname=referral.lastname, student_number=referral.student_number,
    #                                                     degree_program=referral.degree_program,
    #                                                     counselor_id=referral.counselor_id,
    #                                                     start_time=referral.start_time, end_time=referral.end_time,
    #                                                      date=referral.date,
    #                                                     status=referral.status))
                
    # elif (status == 'pending'):
    #     for obj in referrals:
    #         if obj.status == 'pending':
    #             for check in obj.referral_id:
    #                 for obj1 in check:
    #                     get_details = ReferralDetails.objects.get(id = obj1)
    #                     if get_details.faculty_id is not None:
    #                         if user == get_details.faculty_id.faculty_id:
    #                             flag = obj.id
    #             referral = Referral.objects.get(id=flag)
    #             qs.append(Referral(id=flag,firstname=referral.firstname,
    #                                                             lastname=referral.lastname, student_number=referral.student_number,
    #                                                             degree_program=referral.degree_program,
    #                                                             counselor_id=referral.counselor_id,
    #                                                             start_time=referral.start_time, end_time=referral.end_time,
    #                                                             date=referral.date,
    #                                                             status=referral.status))
        
    # elif (status == 'done'):
    #     for obj in referrals:
    #         if obj.status == 'done':
    #             for check in obj.referral_id:
    #                 for obj1 in check:
    #                     get_details = ReferralDetails.objects.get(id = obj1)
    #                     if get_details.faculty_id is not None:
    #                         if user == get_details.faculty_id.faculty_id:
    #                             flag = obj.id
    #             referral = Referral.objects.get(id=flag)
    #             qs.append(Referral(id=flag,firstname=referral.firstname,
    #                                                             lastname=referral.lastname, student_number=referral.student_number,
    #                                                             degree_program=referral.degree_program,
    #                                                             counselor_id=referral.counselor_id,
    #                                                             start_time=referral.start_time, end_time=referral.end_time,
    #                                                             date=referral.date,
    #                                                             status=referral.status))
    filterform = FilterForm()
    if request.method == "POST":
        filterform = FilterForm(request.POST)
        if filterform.is_valid():
            filter_choice = filterform['filter_choice'].value()
            return redirect('teacher_view_referred_students', status=filter_choice)
    return render(request, "teacher/list_referred_students.html", {'teacherNotif': teacherNotif, "status": status, "filterform": filterform, "object_list": qs, "form": teacher_name})


@login_required(login_url='login')
def detail_referred_student(request, id):
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    detail = []
    referrals = Referral.objects.get(id = id)
    for obj in referrals.referral_id:
            get_details = ReferralDetails.objects.get(id=obj)
            if get_details.faculty_id is not None:
                if user == get_details.faculty_id.faculty_id:
                    detail.append(ReferralDetails(subject_referred=get_details.subject_referred,
                                                    reasons=get_details.reasons, behavior_problem=get_details.behavior_problem,
                                                    faculty_id=get_details.faculty_id))
    return render(request, "teacher/detail_referred_student.html", {'teacherNotif': teacherNotif, "detail": detail,'referrals':referrals, "form": teacher_name})


@login_required(login_url='login')
def referral(request, studentReferredId, offer_no):
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)

    # getting the subject of the student being referred
    subject_referred = SubjectOfferings.objects.get(offer_no=offer_no)
    # getting the student being referred
    studentReferred = Student.objects.get(student_number=studentReferredId)
    subject_referred_code = subject_referred.subject_code_id
    referral_form = ReferralForm(initial={'subject_referred': subject_referred_code})

    faculty_load_object = SubjectOfferings.objects.filter(faculty_id=user)
    degree = DegreeProgram.objects.get(
        program_code=studentReferred.program_code_id)
    degree_program_student_referred = degree.program_code

    if request.method == "POST":
        referral_form = ReferralForm(request.POST, initial={ 'subject_referred': subject_referred_code})
        if referral_form.is_valid():
            print('valid')
            today = date.today()
            now = dt.datetime.now()
            classes_counselor = []
            notAvailableSched = []
            sample = []
            tomorrow = today
            finder = 0

            get_object_counselor_assigned = DegreeProgram.objects.get(
                program_code=degree_program_student_referred)
            counselor_assigned_id = get_object_counselor_assigned.faculty_id_id
            CounselorLoad = SubjectOfferings.objects.filter(
                faculty_id=counselor_assigned_id)

            for obj in CounselorLoad:
                print(obj.school_time)

            timeArray = []
            initialtime = 0
            newTime = str(initialtime)+':00:00'

            for x in range(24):
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                newTime = str(initialtime)+':30:00'
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                initialtime = initialtime + 1
                newTime = str(initialtime)+':00:00'

            while(finder == 0):
                tomorrow = tomorrow+timedelta(days=1)
                day_name = tomorrow.strftime("%a")
                if(day_name != "Sun" and day_name != "Sat"):
                    referral_list_byday = Referral.objects.filter(
                        date=tomorrow, counselor_id=counselor_assigned_id)
                    CounselorLoadCheck = bool(CounselorLoad)
                    referral_list_byday_check = bool(referral_list_byday)
                    if(CounselorLoadCheck == True):
                        for object in CounselorLoad:
                            if day_name == 'Thu':
                                check = bool(
                                    day_name[:-1].upper() in object.school_days)
                            else:
                                check = bool(
                                    day_name[0].upper() in object.school_days)
                            if(check == True):
                                classes_counselor.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                                                                          subject_title=object.subject_title, school_days=object.school_days,
                                                                          school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                                                                          department_code=object.department_code, faculty_id=object.faculty_id))
                    else:
                        CounselorLoadCheck = False
                    classes_counselor_check = bool(classes_counselor)

                    notAvailableSched = SetScheduleCounselor.objects.filter(
                        date=tomorrow, faculty_id=counselor_assigned_id)
                    notAvailableSchedChecker = bool(notAvailableSched)

                    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
                    end = datetime.strptime('17:30:00', '%H:%M:%S').time()

                    TimeTaken = 0
                    TimeTaken1 = 0
                    TimeTaken2 = 0
                    counter = 0

                    if (classes_counselor_check == False and referral_list_byday_check == False and notAvailableSchedChecker == False):
                        print('1')
                        startTime = datetime.strptime(
                            '8:00:00', '%H:%M:%S').time()
                        endTime = datetime.strptime(
                            '9:00:00', '%H:%M:%S').time()
                        time1 = startTime
                        time2 = endTime
                        finder = 1
                    elif(classes_counselor_check == True and referral_list_byday_check == False and notAvailableSchedChecker == False):
                        print('2')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0

                    elif classes_counselor_check == False and referral_list_byday_check == True and notAvailableSchedChecker == False:
                        print('3')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in referral_list_byday:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0

                    elif(classes_counselor_check == False and referral_list_byday_check == False and notAvailableSchedChecker == True):
                        print('4')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in notAvailableSched:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0
                    elif(classes_counselor_check == True and referral_list_byday_check == True and notAvailableSchedChecker == False):
                        print('5')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == True and referral_list_byday_check == False and notAvailableSchedChecker == True):
                        print('6')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in notAvailableSched:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == False and referral_list_byday_check == True and notAvailableSchedChecker == True):
                        print('7')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in notAvailableSched:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == True and referral_list_byday_check == True and notAvailableSchedChecker == True):
                        print('8')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0):
                                        for object3 in notAvailableSched:
                                            print("3", timeArray[x])
                                            if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                                TimeTaken2 += 1
                                        if(TimeTaken2 == 0 and counter == 0):
                                            print('time1:', timeArray[x])
                                            time1 = timeArray[x]
                                            counter = 1
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                        elif(TimeTaken2 == 0 and counter == 1):
                                            print('time2:', timeArray[x])
                                            time2 = timeArray[x+1]
                                            finder = 1
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            break
                                        elif TimeTaken2 != 0:
                                            time1 = ''
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            counter = 0
                                        else:
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1-0
                                            time1 = ''
                                    else:
                                        TimeTaken1 = 0
                                        counter = 0
                                        time1 = ''
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                    ScheduledReferralbyDay = []
                    ClassesCounselor = []
                    notAvailableSched = []

                    print('time1', time1)
                    print('time2', time2)
                    if(time1 != '' and time2 != ''):
                        print('9')
                        subject_referred = referral_form.cleaned_data['subject_referred']
                        reasons = referral_form.cleaned_data['reasons']
                        behavior = referral_form.cleaned_data['behavior_problem']
                        faculty = Faculty.objects.get(faculty_id =user)
                        try:
                            check_if_exist = Referral.objects.get(
                                student_number=studentReferred.student_number, date=tomorrow, status='pending')
                            flag = True
                        except Exception:
                            flag = False
                        if flag == True:
                            studentInfo = ReferralDetails(
                                                   subject_referred=subject_referred,
                                                   reasons=reasons,
                                                   faculty_id=faculty,
                                                   behavior_problem=behavior)
                            studentInfo.save() 
                            get_referral_id = ReferralDetails.objects.last()
                            referral_list = [get_referral_id.id]
                            for obj in check_if_exist.referral_id:
                                referral_list.append(obj)
                            check_if_exist.referral_id = referral_list
                            check_if_exist.save()
                        else:
                            studentInfo = ReferralDetails(
                                                   subject_referred=subject_referred,
                                                   reasons=reasons,
                                                   faculty_id=faculty,
                                                   behavior_problem=behavior)
                            studentInfo.save() 
                            referral_id = ReferralDetails.objects.last()
                            studentSched = Referral(student_number=studentReferred.student_number,
                                                   firstname=studentReferred.firstname,
                                                   lastname=studentReferred.lastname,
                                                   middlename=studentReferred.middlename,
                                                   degree_program=studentReferred.program_code_id,
                                                   counselor_id=counselor_assigned_id,
                                                   start_time=time1, end_time=time2, date=tomorrow,
                                                   choice='Counseling',referral_id = [referral_id.id])
                            studentSched.save()
                            create_notification(counselor_assigned_id, user, 'manual_referral', extra_id=int(
                                studentReferred.student_number), schedDay=tomorrow, schedStartTime=time1, schedEndTime=time2)
                        referral_form = ReferralForm(request.POST, initial={ 'subject_referred': subject_referred_code})
                        messages.info(
                            request, 'Successfully Referred the Student')
    return render(request, "teacher/refer_a_student.html", {'teacherNotif': teacherNotif, "referral_form": referral_form, "form": teacher_name})


@login_required
def teacher_notifications(request):
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = NotificationFeedback.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == NotificationFeedback.AUTOMATIC_REFERRAL:
            return render(request, "teacher/counselor.html", {})
        elif notification.notification_type == NotificationFeedback.MANUAL_REFERRAL:
            return render(request, "teacher/counselor.html", {})

    counselorNotif = NotificationFeedback.objects.filter(
        to_user=user, notification_type="feedback_teacher").order_by('created_at')
    return render(request, 'teacher/notification.html', {"notifications": counselorNotif, "form": teacher_name})


@login_required(login_url='login')
def teacher_notification_detail(request, id):
    user = request.session.get('username')
    teacher_name = Faculty.objects.get(faculty_id=user)
    notif = NotificationFeedback.objects.filter(to_user = user , is_read=False)
    teacherNotif = len(notif)
    detail = []
    notification = NotificationFeedback.objects.get(id=id)
    notification.is_read = True
    notification.save()
    student = Referral.objects.get(id=notification.extra_id)
    for obj in student.referral_id:
            get_details = ReferralDetails.objects.get(id=obj)
            if get_details.faculty_id is not None:
                if user == get_details.faculty_id.faculty_id:
                    detail.append(ReferralDetails(subject_referred=get_details.subject_referred,
                                                    reasons=get_details.reasons, behavior_problem=get_details.behavior_problem,
                                                    faculty_id=get_details.faculty_id))
    return render(request, "teacher/teacher_notification_detail.html", {'teacherNotif': teacherNotif, "detail": detail,'student':student, "form": teacher_name})

# teacher

# student


@login_required(login_url='login')
def student_add_information(request, *args, **kwargs):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    infoForm = StudentAdditionalInformationForm(instance=student_name)
    if request.method == "POST":
        infoForm = StudentAdditionalInformationForm(
            request.POST, instance=student_name)
        print(infoForm.errors)
        if infoForm.is_valid():
            student_number = infoForm['student_number'].value()
            lastname = infoForm['lastname'].value()
            firstname = infoForm['firstname'].value()
            middlename = infoForm['middlename'].value()
            student_email = infoForm['student_email'].value()
            student_contact_number = infoForm['student_contact_number'].value()
            mother_firstname = infoForm['mother_firstname'].value()
            mother_lastname = infoForm['mother_lastname'].value()
            father_firstname = infoForm['father_firstname'].value()
            father_lastname = infoForm['father_lastname'].value()
            guardian_firstname = infoForm['guardian_firstname'].value()
            guardian_lastname = infoForm['guardian_lastname'].value()
            student_contact_number = infoForm['student_contact_number'].value()
            mother_contact_number = infoForm['mother_contact_number'].value()
            father_contact_number = infoForm['father_contact_number'].value()
            guardian_contact_number = infoForm['guardian_contact_number'].value(
            )
            info = StudentAdditionalInformation(student_number=student_number, lastname=lastname,
                                                firstname=firstname, middlename=middlename, student_email=student_email, student_contact_number=student_contact_number,
                                                mother_firstname=mother_firstname, mother_lastname=mother_lastname, mother_contact_number=mother_contact_number,
                                                father_firstname=father_firstname, father_lastname=father_lastname, father_contact_number=father_contact_number,
                                                guardian_firstname=guardian_firstname, guardian_lastname=guardian_lastname, guardian_contact_number=guardian_contact_number,
                                                status="done")
            info.save()
            return redirect('student_home_view')
    return render(request, "student/add_information.html", {"form": student_name, "info": infoForm})


@login_required(login_url='login')
def student_home_view(request, *args, **kwargs):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    return render(request, "student/home.html", {"studentNotif": studentNotif, "form": student_name})


@login_required
def student_notifications(request):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    goto = request.GET.get('goto', '')
    notification_id = request.GET.get('notification', 0)
    extra_id = request.GET.get('extra_id', 0)

    if goto != '':
        notification = Notification.objects.get(pk=notification_id)
        notification.is_read = True
        notification.save()

        if notification.notification_type == Notification.AUTOMATIC_REFERRAL:
            return render(request, "student/student_home.html", {})
        elif notification.notification_type == Notification.MANUAL_REFERRAL:
            return render(request, "student/student_home", {})
    notif = Notification.objects.filter(extra_id=user).order_by('created_at')
    return render(request, 'student/notification.html', {"studentNotif": studentNotif, "notifications": notif, "form": student_name})


@login_required
def student_notification_detail(request, id):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    notification = Notification.objects.get(id=id)
    notification.is_read_student = True
    notification.save()
    detail = []
    student = Referral.objects.get(id=id)
    return render(request, 'student/student_notification_detail.html', {"studentNotif": studentNotif, 'student':student, "form": student_name})


@login_required(login_url='login')
def edit_information(request, *args, **kwargs):
    user = request.session.get('username')
    student_name = StudentAdditionalInformation.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    infoForm = StudentAdditionalInformationForm(instance=student_name)
    if request.method == "POST":
        infoForm = StudentAdditionalInformationForm(
            request.POST, instance=student_name)
        if infoForm.is_valid():
            student_number = infoForm['student_number'].value()
            lastname = infoForm['lastname'].value()
            firstname = infoForm['firstname'].value()
            middlename = infoForm['middlename'].value()
            student_email = infoForm['student_email'].value()
            student_contact_number = infoForm['student_contact_number'].value()
            mother_firstname = infoForm['mother_firstname'].value()
            mother_lastname = infoForm['mother_lastname'].value()
            father_firstname = infoForm['father_firstname'].value()
            father_lastname = infoForm['father_lastname'].value()
            guardian_firstname = infoForm['guardian_firstname'].value()
            guardian_lastname = infoForm['guardian_lastname'].value()
            student_contact_number = infoForm['student_contact_number'].value()
            mother_contact_number = infoForm['mother_contact_number'].value()
            father_contact_number = infoForm['father_contact_number'].value()
            guardian_contact_number = infoForm['guardian_contact_number'].value(
            )
            info = StudentAdditionalInformation(student_number=student_number, lastname=lastname,
                                                firstname=firstname, middlename=middlename, student_email=student_email, student_contact_number=student_contact_number,
                                                mother_firstname=mother_firstname, mother_lastname=mother_lastname, mother_contact_number=mother_contact_number,
                                                father_firstname=father_firstname, father_lastname=father_lastname, father_contact_number=father_contact_number,
                                                guardian_firstname=guardian_firstname, guardian_lastname=guardian_lastname, guardian_contact_number=guardian_contact_number,
                                                status="done")
            info.save()
            messages.info(request, 'Success')
    return render(request, "student/edit_information.html", {"studentNotif": studentNotif, "form": student_name, "info": infoForm})


@login_required
def student_history(request):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    student_record = Referral.objects.filter(
        student_number=user, status='done')
    counselor = Faculty.objects.filter(role='Counselor')
    return render(request, "student/view_history.html", {"studentNotif": studentNotif, "object": student_record, "counselor": counselor, "form": student_name})


@login_required
def student_view_schedule(request, *args, **kwargs):
    global Active_Sem
    global Active_Year
    global count1
    count1 = 0
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    today = date.today()
    now = dt.datetime.now()
    day_name = now.strftime("%a")

    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        print('hello')
        print(offer.errors)
        if offer.is_valid():
            print('hi')
            offer.save()

    classes_of_student = []
    ScheduledReferralbyDayCheck = False

    classes_of_student_list = Studentload.objects.filter(student_number=user ,sem_id = Active_Sem, academic_year=Active_Year)
    classes_of_student_list_checker = bool(classes_of_student_list)

    if(classes_of_student_list_checker == True):
        for object in classes_of_student_list:
            print('object.offer_no.offer_no',object.offer_no.offer_no)
            try:
                get_school_days = SubjectOfferings.objects.get(
                    offer_no=object.offer_no.offer_no, sem_id = Active_Sem, academic_year=Active_Year)
            except:
                check=False
            if day_name == 'Thu':
                check = bool(
                    day_name[:-1].upper() in get_school_days.school_days)
            else:
                check = bool(
                    day_name[0].upper() in get_school_days.school_days)
            if(check == True):
                classes_of_student.append(SubjectOfferings(offer_no=get_school_days.offer_no, subject_code=get_school_days.subject_code,
                                                           subject_title=get_school_days.subject_title, school_days=get_school_days.school_days,
                                                           school_time=get_school_days.school_time, sem_id=get_school_days.sem_id, academic_year=get_school_days.academic_year,
                                                           department_code=get_school_days.department_code, faculty_id=get_school_days.faculty_id))

    referral_list_byday = Referral.objects.filter(
        student_number=user, date=today).order_by('start_time')
    referral_list_byday_checker = bool(referral_list_byday)
    classes_of_student_checker = bool(classes_of_student)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    sched_for_today = []

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
    end = datetime.strptime('17:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    initialtime = 7
    id = 0
    for x in range(12):
            id +=1
            newTime = str(initialtime)+':00:00'
            time1 = datetime.strptime(newTime, '%H:%M:%S').time()
            newTime = str(initialtime)+':30:00'
            time2 = datetime.strptime(newTime, '%H:%M:%S').time()
            initialtime += 1
            time = NewTime(time_id=id, time1=time1, time2=time2)
            time.save()
            flag=1
            while(flag==1):
                id +=1
                newTime = str(initialtime)+':00:00'
                next_time = datetime.strptime(newTime, '%H:%M:%S').time()
                time = NewTime(time_id=id, time1=time2, time2=next_time)
                time.save()
                flag=0
    alltime = NewTime.objects.all()

    if(referral_list_byday_checker == True):
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    sched_for_today.append(Referral(id= object.id,student_number=object.student_number, firstname=object.firstname,
                                                    lastname=object.lastname, middlename=object.middlename,
                                                    degree_program=object.degree_program, counselor_id=object.counselor_id, status=object.status, 
                                                    start_time=object.start_time, end_time=object.end_time, date=object.date, 
                                                    choice=choice))
    else:
        print('no display')
    return render(request, "student/student_view_schedule.html", {"offer": offer, "studentNotif": studentNotif, "today": today, "day_name": day_name, "schedForToday": sched_for_today, "time": alltime, "form": student_name})


@login_required
def another_student_view_schedule(request, *args, **kwargs):
    global count1
    global Active_Sem
    global Active_Year
    count1 = 0
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    
    offer = CalendarForm()
    if request.method == "POST":
        offer = CalendarForm(request.POST)
        if offer.is_valid():
            offer.save()

    newDate = Calendar.objects.last()
    today = newDate.pickedDate
    day_name = today.strftime("%a")

    classes_of_student = []
    ScheduledReferralbyDayCheck = False

    classes_of_student_list = Studentload.objects.filter(student_number=user)
    classes_of_student_list_checker = bool(classes_of_student_list)

    if(classes_of_student_list_checker == True):
        for object in classes_of_student_list:
            try:
                get_school_days = SubjectOfferings.objects.get(
                    offer_no=object.offer_no.offer_no, sem_id = Active_Sem, academic_year=Active_Year)
            except:
                check=False
            if day_name == 'Thu':
                check = bool(
                    day_name[:-1].upper() in get_school_days.school_days)
            else:
                check = bool(
                    day_name[0].upper() in get_school_days.school_days)
            if(check == True):
                classes_of_student.append(SubjectOfferings(offer_no=get_school_days.offer_no, subject_code=get_school_days.subject_code,
                                                           subject_title=get_school_days.subject_title, school_days=get_school_days.school_days,
                                                           school_time=get_school_days.school_time, sem_id=get_school_days.sem_id, academic_year=get_school_days.academic_year,
                                                           department_code=get_school_days.department_code, faculty_id=get_school_days.faculty_id))

    referral_list_byday = Referral.objects.filter(
        student_number=user, date=today).order_by('start_time')
    referral_list_byday_checker = bool(referral_list_byday)
    classes_of_student_checker = bool(classes_of_student)

    timeArray = []
    initialtime = 0

    newTime = str(initialtime)+':00:00'
    sched_for_today = []

    for x in range(24):
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        newTime = str(initialtime)+':30:00'
        timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
        initialtime = initialtime + 1
        newTime = str(initialtime)+':00:00'

    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
    end = datetime.strptime('17:00:00', '%H:%M:%S').time()
    check = datetime.strptime('00:00:00', '%H:%M:%S').time()
    initialtime = 7
    id = 0
    for x in range(12):
            id +=1
            newTime = str(initialtime)+':00:00'
            time1 = datetime.strptime(newTime, '%H:%M:%S').time()
            newTime = str(initialtime)+':30:00'
            time2 = datetime.strptime(newTime, '%H:%M:%S').time()
            initialtime += 1
            time = NewTime(time_id=id, time1=time1, time2=time2)
            time.save()
            flag=1
            while(flag==1):
                id +=1
                newTime = str(initialtime)+':00:00'
                next_time = datetime.strptime(newTime, '%H:%M:%S').time()
                time = NewTime(time_id=id, time1=time2, time2=next_time)
                time.save()
                flag=0
    alltime = NewTime.objects.all()

    if(referral_list_byday_checker == True):
        for x in range(len(timeArray)):
            for object in referral_list_byday:
                if(timeArray[x] == object.start_time):
                    choice = 'Counseling'
                    sched_for_today.append(Referral(id=object.id,student_number=object. student_number, firstname=object.firstname,
                                                    lastname=object.lastname, middlename=object.middlename,
                                                    degree_program=object.degree_program, 
                                                     counselor_id=object.counselor_id, status=object.status, 
                                                    start_time=object.start_time, end_time=object.end_time, date=object.date,
                                                    choice=choice))
    else:
        print('no display')
    return render(request, "student/student_view_schedule.html", {"offer": offer, "studentNotif": studentNotif, "today": today, "day_name": day_name, "schedForToday": sched_for_today, "time": alltime, "form": student_name})


@login_required(login_url='login')
def student_set_schedule(request, *args, **kwargs):
    user = request.session.get('username')
    student_name = Student.objects.get(student_number=user)
    notif = Notification.objects.filter(extra_id=user, is_read_student=False)
    studentNotif = len(notif)
    schedForm = StudentSetSchedForm()
    studentReferred = Student.objects.get(student_number=user)
    degree = DegreeProgram.objects.get(
        program_code=studentReferred.program_code_id)
    degree_program_student_referred = degree.program_code
    if request.method == "POST":
        schedForm = StudentSetSchedForm(request.POST)
        print(schedForm.errors)
        if schedForm.is_valid():
            today = date.today()
            now = dt.datetime.now()
            classes_counselor = []
            notAvailableSched = []
            sample = []
            tomorrow = today
            finder = 0

            get_object_counselor_assigned = DegreeProgram.objects.get(
                program_code=degree_program_student_referred)
            counselor_assigned_id = get_object_counselor_assigned.faculty_id_id
            CounselorLoad = SubjectOfferings.objects.filter(
                faculty_id=counselor_assigned_id)
            timeArray = []
            initialtime = 0
            newTime = str(initialtime)+':00:00'

            for x in range(24):
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                newTime = str(initialtime)+':30:00'
                timeArray.append(datetime.strptime(newTime, '%H:%M:%S').time())
                initialtime = initialtime + 1
                newTime = str(initialtime)+':00:00'

            while(finder == 0):
                tomorrow = tomorrow+timedelta(days=1)
                day_name = tomorrow.strftime("%a")
                if(day_name != "Sun" and day_name != "Sat"):
                    referral_list_byday = Referral.objects.filter(
                        date=tomorrow)
                    CounselorLoadCheck = bool(CounselorLoad)
                    referral_list_byday_check = bool(referral_list_byday)
                    if(CounselorLoadCheck == True):
                        for object in CounselorLoad:
                            if day_name == 'Thu':
                                check = bool(
                                    day_name[:-1].upper() in object.school_days)
                            else:
                                check = bool(
                                    day_name[0].upper() in object.school_days)
                            if(check == True):
                                classes_counselor.append(SubjectOfferings(offer_no=object.offer_no, subject_code=object.subject_code,
                                                                          subject_title=object.subject_title, school_days=object.school_days,
                                                                          school_time=object.school_time, sem_id=object.sem_id, academic_year=object.academic_year,
                                                                          department_code=object.department_code, faculty_id=object.faculty_id))
                    else:
                        CounselorLoadCheck = False
                    classes_counselor_check = bool(classes_counselor)

                    notAvailableSched = SetScheduleCounselor.objects.filter(
                        date=tomorrow, faculty_id=counselor_assigned_id)
                    notAvailableSchedChecker = bool(notAvailableSched)

                    start = datetime.strptime('8:00:00', '%H:%M:%S').time()
                    end = datetime.strptime('17:30:00', '%H:%M:%S').time()

                    TimeTaken = 0
                    TimeTaken1 = 0
                    TimeTaken2 = 0
                    counter = 0

                    if (classes_counselor_check == False and referral_list_byday_check == False and notAvailableSchedChecker == False):
                        print('1')
                        startTime = datetime.strptime(
                            '8:00:00', '%H:%M:%S').time()
                        endTime = datetime.strptime(
                            '9:00:00', '%H:%M:%S').time()
                        time1 = startTime
                        time2 = endTime
                        finder = 1
                    elif(classes_counselor_check == True and referral_list_byday_check == False and notAvailableSchedChecker == False):
                        print('2')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0

                    elif classes_counselor_check == False and referral_list_byday_check == True and notAvailableSchedChecker == False:
                        print('3')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in referral_list_byday:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0

                    elif(classes_counselor_check == False and referral_list_byday_check == False and notAvailableSchedChecker == True):
                        print('4')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in notAvailableSched:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0 and counter == 0):
                                    time1 = timeArray[x]
                                    counter = 1
                                    TimeTaken = 0
                                elif(TimeTaken == 0 and counter == 1):
                                    time2 = timeArray[x+1]
                                    counter = 0
                                    finder = 1
                                    break
                                elif(TimeTaken != 0):
                                    time1 = ''
                                    counter = 0
                                    TimeTaken = 0
                    elif(classes_counselor_check == True and referral_list_byday_check == True and notAvailableSchedChecker == False):
                        print('5')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == True and referral_list_byday_check == False and notAvailableSchedChecker == True):
                        print('6')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in notAvailableSched:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == False and referral_list_byday_check == True and notAvailableSchedChecker == True):
                        print('7')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in notAvailableSched:
                                    if(timeArray[x+1] <= object.end_time and timeArray[x] >= object.start_time):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0 and counter == 0):
                                        time1 = timeArray[x]
                                        counter = 1
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                    elif (TimeTaken1 == 0 and counter == 1):
                                        time2 = timeArray[x+1]
                                        finder = 1
                                        counter = 0
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        break
                                    elif(TimeTaken1 != 0):
                                        time1 = ''
                                        TimeTaken = 0
                                        TimeTaken1 = 0
                                        counter = 0

                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                            else:
                                counter = 0
                                TimeTaken = 0
                                time1 = ''

                    elif(classes_counselor_check == True and referral_list_byday_check == True and notAvailableSchedChecker == True):
                        print('8')
                        for x in range(len(timeArray)):
                            if(timeArray[x] >= start and timeArray[x] < end):
                                for object in classes_counselor:
                                    get_time = object.school_time
                                    classes_counselor_time = get_time.split(
                                        '-')
                                    classes_counselor_start_time = classes_counselor_time[0].upper(
                                    ).replace(" ", "")
                                    classes_counselor_end_time = classes_counselor_time[1].upper(
                                    ).replace(" ", "")

                                    # for start_time
                                    if (classes_counselor_start_time[1]) == ':':
                                        classes_counselor_start_time = "".join(
                                            ('0', classes_counselor_start_time))
                                    if classes_counselor_start_time[-2:] == "AM":
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = str(
                                                '00' + classes_counselor_start_time[2:-2])
                                        else:
                                            cc_start = classes_counselor_start_time[:-2]
                                    else:
                                        if classes_counselor_start_time[:2] == '12':
                                            cc_start = classes_counselor_start_time[:-2]
                                        else:
                                            cc_start = str(
                                                int(classes_counselor_start_time[:2]) + 12) + classes_counselor_start_time[2:-2]
                                    # for end_time
                                    if classes_counselor_end_time[1] == ':':
                                        classes_counselor_end_time = "".join(
                                            ('0', classes_counselor_end_time))
                                    if classes_counselor_end_time[-2:] == "AM":
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = str(
                                                '00' + classes_counselor_end_time[2:-2])
                                        else:
                                            cc_end = classes_counselor_end_time[:-2]
                                    else:
                                        if classes_counselor_end_time[:2] == '12':
                                            cc_end = classes_counselor_end_time[:-2]
                                        else:
                                            cc_end = str(
                                                int(classes_counselor_end_time[:2]) + 12) + classes_counselor_end_time[2:-2]
                                    cc_start_convert = datetime.strptime(
                                        cc_start, '%H:%M').time()
                                    cc_end_convert = datetime.strptime(
                                        cc_end, '%H:%M').time()
                                    if(timeArray[x+1] <= cc_end_convert and timeArray[x] >= cc_start_convert):
                                        TimeTaken += 1
                                if(TimeTaken == 0):
                                    for object2 in referral_list_byday:
                                        if(timeArray[x+1] <= object2.end_time and timeArray[x] >= object2.start_time):
                                            TimeTaken1 += 1
                                    if(TimeTaken1 == 0):
                                        for object3 in notAvailableSched:
                                            print("3", timeArray[x])
                                            if(timeArray[x+1] <= object3.end_time and timeArray[x] >= object3.start_time):
                                                TimeTaken2 += 1
                                        if(TimeTaken2 == 0 and counter == 0):
                                            print('time1:', timeArray[x])
                                            time1 = timeArray[x]
                                            counter = 1
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                        elif(TimeTaken2 == 0 and counter == 1):
                                            print('time2:', timeArray[x])
                                            time2 = timeArray[x+1]
                                            finder = 1
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            break
                                        elif TimeTaken2 != 0:
                                            time1 = ''
                                            TimeTaken = 0
                                            TimeTaken1 = 0
                                            TimeTaken2 = 0
                                            counter = 0
                                        else:
                                            counter = 0
                                            TimeTaken = 0
                                            TimeTaken1-0
                                            time1 = ''
                                    else:
                                        TimeTaken1 = 0
                                        counter = 0
                                        time1 = ''
                                else:
                                    counter = 0
                                    TimeTaken = 0
                                    time1 = ''
                    ScheduledReferralbyDay = []
                    ClassesCounselor = []
                    notAvailableSched = []

                    print('time1', time1)
                    print('time2', time2)
                    if(time1 != '' and time2 != ''):
                        print('9')
                        list_reasons = []
                        reasons = schedForm.cleaned_data['reasons']
                        try:
                            check_if_exist = Referral.objects.get(
                                student_number=studentReferred.student_number, date=tomorrow, status='pending')
                            flag = True
                        except Exception:
                            flag = False
                        if flag == False:
                            studentInfo = ReferralDetails(reasons=reasons)
                            studentInfo.save() 
                            referral_id = ReferralDetails.objects.last()
                            studentSched = Referral(student_number=studentReferred.student_number,
                                                   firstname=studentReferred.firstname,
                                                   lastname=studentReferred.lastname,
                                                   middlename=studentReferred.middlename,
                                                   degree_program=studentReferred.program_code_id,
                                                   counselor_id=counselor_assigned_id,
                                                   start_time=time1, end_time=time2, date=tomorrow,
                                                   choice='Counseling',referral_id = [referral_id.id])
                            studentSched.save()
                            create_notification(counselor_assigned_id, user, 'manual_referral', extra_id=int(
                                studentReferred.student_number), schedDay=tomorrow, schedStartTime=time1, schedEndTime=time2)
                        else:
                            studentInfo = ReferralDetails(reasons=reasons)
                            studentInfo.save() 
                            get_referral_id = ReferralDetails.objects.last()
                            referral_list = [get_referral_id.id]
                            for obj in check_if_exist.referral_id:
                                referral_list.append(obj)
                            check_if_exist.referral_id = referral_list
                            check_if_exist.save()
                        schedForm = StudentSetSchedForm()
                        messages.info(
                            request, 'Successfully Set Schedule')
    return render(request, "student/set_schedule.html", {"studentNotif": studentNotif, "schedform": schedForm, "form": student_name})

# student



# try:
    #     a = ReferralDetails.objects.get(id = 1)
    # except:
    #     print('way sud')
    # today = date.today()
    # day_name = today.strftime("%a")
    # print(day_name[:-1])
    # Referral.objects.all().delete()
    # Notification.objects.all().delete()
    # timeArray = []
    # newTime = str('2')+':00'
    # timeArray.append(datetime.strptime(newTime, '%H:%M').time())
    # newTime = str('2')+':30'
    # d = datetime.strptime(newTime, "%H:%M").time()
    # with_ampm = d.strftime("%I:%M %p")
    # print('with_ampm', with_ampm)
    # print('d', d)
    # print('newTime', newTime)
    # print('timeArray', timeArray)
    # subject = SubjectOfferings.objects.get(offer_no='60002')
    # splitDate = subject.school_time.split('-')
    # check_a = 'hello'
    # check_b = 'l'
    # if check_b in check_a:
    #     print('work')
    # checkaaa = bool(check_b in check_a)
    # print(checkaaa)
    # time = '12:30 pm'
    # convert_into_time = datetime.strptime(time, '%H:%M %p').time()
    # convert_into_time_with_ampm = convert_into_time.strftime('%I:%M %p')
    # print(convert_into_time, type(convert_into_time))
    # print(convert_into_time_with_ampm, type(convert_into_time_with_ampm))
    # time1 = '8:30 pm'
    # time2 = '8:30 pm'

    # print(time2, type(time2))
    # d = datetime.strptime("22:30", "%H:%M")
    # with_ampm = d.strftime("%I:%M %p")
    # print(d.strftime("%I:%M %p"))

    # today = date.today()
    # day_name = today.strftime("%a")
    # abbreviation = day_name[0]
    # print('abbreviation', abbreviation)
    # print('day_name[0]', day_name[0])
    # today = date.today()
    # day_name = today.strftime("%a")
    # abbreviation = day_name[0:2]
    # print(abbreviation)
    # print(day_name[0:2])
    # a = '10124534534534534454'
    # ap = []
    # get = []
    # value = ChessBoard.objects.get(id=2)
    # print(value.board)
    # for obj in value.board:
    #     get.append(obj)
    # get.append(a)
    # print(get)
    # value.board = get
    # value.save()
    # print(value.board)